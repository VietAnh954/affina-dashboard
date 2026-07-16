"""
================================================================================
 EXPORT FULL EXCEL — Xuất toàn bộ dữ liệu ra 1 file Excel với 10 sheet phân tích
================================================================================

Cách chạy:
    1. Đảm bảo có file .env với SUPABASE_DB_URI (hoặc set env variable)
    2. Cài dependencies:
         pip install pandas sqlalchemy psycopg2-binary openpyxl python-dotenv
    3. Chạy:
         python scripts/export_full_excel.py
    4. File Excel sẽ được tạo tại: ./affina_full_export_YYYYMMDD_HHMM.xlsx

File output có 10 sheet:
  📊 Summary        — KPI tổng hợp
  📁 All_Data       — Full raw data (dashboard_master_data)
  🎨 By_Source      — Pivot theo Core/Neo/TSA
  📡 By_Channel     — Pivot theo Channel
  🛡 By_Loai_BH     — Pivot theo Loại bảo hiểm
  🏢 By_Nha_BH      — Pivot theo Nhà bảo hiểm
  📦 Top_Products   — Top 30 sản phẩm
  👥 Sales_Ranking  — Ranking toàn bộ sale
  📅 By_Month       — Time series theo tháng
  ℹ️ Metadata       — Info snapshot

Author: Senior DA companion for VietAnh @ Affina
================================================================================
"""
import os
import sys
from datetime import datetime, timezone, timedelta

import pandas as pd
from sqlalchemy import create_engine

# .env cho local dev
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# openpyxl cho format
try:
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    print("❌ Missing openpyxl. Install: pip install openpyxl")
    sys.exit(1)


VN_TZ = timezone(timedelta(hours=7))


def log(msg: str) -> None:
    ts = datetime.now(VN_TZ).strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


CURRENCY_COLS = {
    "Số tiền thanh toán", "Phí BH (VNĐ)", "Doanh thu trước thuế",
    "EST_Bonus", "Affina_Revenue", "Thưởng Teamlead", "Incentive OVE",
    "SM_OR", "SD_OR", "SM_IO", "SD_IO", "BDM_bonus", "BDD_bonus",
    "Chi Agency", "Chi QL", "Budget Neo T6",
    "total_revenue", "revenue", "affina", "bonus",
}
DATE_COLS_KEYWORDS = ["Ngày", "date", "month"]


def _strip_tz(df: pd.DataFrame) -> pd.DataFrame:
    """Bỏ timezone khỏi mọi cột datetime — Excel không hỗ trợ tz-aware datetime.
    Cột _ingested_at từ Supabase (TIMESTAMPTZ) gây lỗi nếu không strip."""
    df = df.copy()
    for c in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[c]):
            try:
                if getattr(df[c].dt, "tz", None) is not None:
                    df[c] = df[c].dt.tz_localize(None)
            except (TypeError, AttributeError):
                pass
    return df


def _apply_format(ws, df, header_color="305496"):
    """Format 1 sheet — header, filter, freeze, auto-width, number format."""
    if ws.max_row < 1:
        return

    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Data cells
    for i, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(i)
        # Number format
        if col_name in CURRENCY_COLS:
            fmt = '#,##0" ₫"'
        elif any(k in str(col_name) for k in DATE_COLS_KEYWORDS):
            fmt = "dd/mm/yyyy"
        else:
            fmt = None

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            cell.border = border
            if fmt:
                cell.number_format = fmt

        # Auto width
        try:
            sample_lens = [len(str(v)) for v in df[col_name].head(50).astype(str).tolist()]
            max_len = max([len(str(col_name))] + sample_lens)
            ws.column_dimensions[col_letter].width = min(35, max(10, max_len + 2))
        except Exception:
            ws.column_dimensions[col_letter].width = 15

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32


def build_summary_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Build KPI summary."""
    rows = []
    total_rev = df["Doanh thu trước thuế"].sum() if "Doanh thu trước thuế" in df else 0
    n_hd = df["Số hợp đồng"].nunique() if "Số hợp đồng" in df else 0
    n_sale = df["Họ tên sale"].nunique() if "Họ tên sale" in df else 0
    total_affina = df["Affina_Revenue"].sum() if "Affina_Revenue" in df else 0
    total_bonus = df["EST_Bonus"].sum() if "EST_Bonus" in df else 0

    rows.append(("📅 Snapshot generated", datetime.now(VN_TZ).strftime("%d/%m/%Y %H:%M:%S")))
    rows.append(("📊 Tổng số dòng", f"{len(df):,}"))
    if "Ngày thanh toán" in df.columns and df["Ngày thanh toán"].notna().any():
        rows.append(("📆 Ngày sớm nhất", df["Ngày thanh toán"].min().strftime("%d/%m/%Y")))
        rows.append(("📆 Ngày muộn nhất", df["Ngày thanh toán"].max().strftime("%d/%m/%Y")))
    rows.append(("", ""))
    rows.append(("─── KEY METRICS ───", ""))
    rows.append(("💰 Tổng doanh thu trước thuế", f"{total_rev:,.0f} ₫"))
    rows.append(("💎 Tổng phí BH", f"{df['Phí BH (VNĐ)'].sum():,.0f} ₫" if 'Phí BH (VNĐ)' in df else "N/A"))
    rows.append(("🏢 Tổng Affina Revenue", f"{total_affina:,.0f} ₫"))
    rows.append(("🎯 Tổng EST Bonus", f"{total_bonus:,.0f} ₫"))
    rows.append(("📋 Số HĐ unique", f"{n_hd:,}"))
    rows.append(("👥 Số Sale unique", f"{n_sale:,}"))
    if n_hd > 0:
        rows.append(("📊 AVG doanh thu / HĐ", f"{total_rev/n_hd:,.0f} ₫"))
    if n_sale > 0:
        rows.append(("📊 AVG doanh thu / Sale", f"{total_rev/n_sale:,.0f} ₫"))

    # Breakdown by Source
    if "Source" in df.columns:
        rows.append(("", ""))
        rows.append(("─── BREAKDOWN BY SOURCE ───", ""))
        for src, grp in df.groupby("Source"):
            rev = grp["Doanh thu trước thuế"].sum() if "Doanh thu trước thuế" in grp else 0
            pct = rev / total_rev * 100 if total_rev > 0 else 0
            rows.append((
                f"  🎨 {src}",
                f"{rev:,.0f} ₫  ({pct:.1f}%, {len(grp):,} dòng)"
            ))

    # Breakdown by Loại BH
    if "Loại bảo hiểm" in df.columns:
        rows.append(("", ""))
        rows.append(("─── BREAKDOWN BY LOẠI BH ───", ""))
        for lb, grp in df.groupby("Loại bảo hiểm"):
            rev = grp["Doanh thu trước thuế"].sum() if "Doanh thu trước thuế" in grp else 0
            pct = rev / total_rev * 100 if total_rev > 0 else 0
            rows.append((f"  🛡 {lb}", f"{rev:,.0f} ₫  ({pct:.1f}%)"))

    return pd.DataFrame(rows, columns=["Chỉ số", "Giá trị"])


def build_by_source(df: pd.DataFrame) -> pd.DataFrame:
    if "Source" not in df.columns:
        return pd.DataFrame()
    return df.groupby("Source", as_index=False).agg(
        so_hd=("Số hợp đồng", "nunique"),
        so_sale=("Họ tên sale", "nunique"),
        doanh_thu=("Doanh thu trước thuế", "sum"),
        phi_bh=("Phí BH (VNĐ)", "sum"),
        affina=("Affina_Revenue", "sum"),
        bonus=("EST_Bonus", "sum"),
    ).sort_values("doanh_thu", ascending=False).rename(columns={
        "so_hd": "Số HĐ", "so_sale": "Số Sale",
        "doanh_thu": "Doanh thu trước thuế", "phi_bh": "Phí BH (VNĐ)",
        "affina": "Affina_Revenue", "bonus": "EST_Bonus",
    })


def build_by_channel(df: pd.DataFrame) -> pd.DataFrame:
    if "Channel" not in df.columns:
        return pd.DataFrame()
    return df.groupby("Channel", as_index=False).agg(
        so_hd=("Số hợp đồng", "nunique"),
        so_sale=("Họ tên sale", "nunique"),
        doanh_thu=("Doanh thu trước thuế", "sum"),
        affina=("Affina_Revenue", "sum"),
        bonus=("EST_Bonus", "sum"),
    ).sort_values("doanh_thu", ascending=False).rename(columns={
        "so_hd": "Số HĐ", "so_sale": "Số Sale",
        "doanh_thu": "Doanh thu trước thuế",
        "affina": "Affina_Revenue", "bonus": "EST_Bonus",
    })


def build_by_loai_bh(df: pd.DataFrame) -> pd.DataFrame:
    if "Loại bảo hiểm" not in df.columns:
        return pd.DataFrame()
    return df.groupby("Loại bảo hiểm", as_index=False).agg(
        so_hd=("Số hợp đồng", "nunique"),
        doanh_thu=("Doanh thu trước thuế", "sum"),
        phi_bh=("Phí BH (VNĐ)", "sum"),
        affina=("Affina_Revenue", "sum"),
    ).sort_values("doanh_thu", ascending=False).rename(columns={
        "so_hd": "Số HĐ", "doanh_thu": "Doanh thu trước thuế",
        "phi_bh": "Phí BH (VNĐ)", "affina": "Affina_Revenue",
    })


def build_by_nha_bh(df: pd.DataFrame) -> pd.DataFrame:
    if "Nhà BH" not in df.columns:
        return pd.DataFrame()
    return df.groupby("Nhà BH", as_index=False).agg(
        so_hd=("Số hợp đồng", "nunique"),
        doanh_thu=("Doanh thu trước thuế", "sum"),
        affina=("Affina_Revenue", "sum"),
    ).sort_values("doanh_thu", ascending=False).rename(columns={
        "so_hd": "Số HĐ", "doanh_thu": "Doanh thu trước thuế",
        "affina": "Affina_Revenue",
    })


def build_top_products(df: pd.DataFrame, top_n: int = 30) -> pd.DataFrame:
    if "Sản phẩm" not in df.columns:
        return pd.DataFrame()
    return df.groupby("Sản phẩm", as_index=False).agg(
        so_hd=("Số hợp đồng", "nunique"),
        doanh_thu=("Doanh thu trước thuế", "sum"),
    ).sort_values("doanh_thu", ascending=False).head(top_n).rename(columns={
        "so_hd": "Số HĐ", "doanh_thu": "Doanh thu trước thuế",
    })


def build_sales_ranking(df: pd.DataFrame) -> pd.DataFrame:
    if "Họ tên sale" not in df.columns:
        return pd.DataFrame()
    result = df.groupby("Họ tên sale", as_index=False).agg(
        chuc_danh=("Chức danh", "first") if "Chức danh" in df.columns else ("Họ tên sale", "count"),
        source=("Source", "first") if "Source" in df.columns else ("Họ tên sale", "count"),
        channel=("Channel", "first") if "Channel" in df.columns else ("Họ tên sale", "count"),
        bdm=("QUẢN LÝ CẤP 1 (BDM)", "first") if "QUẢN LÝ CẤP 1 (BDM)" in df.columns else ("Họ tên sale", "count"),
        bdd=("QUẢN LÝ CẤP 2 (BDD)", "first") if "QUẢN LÝ CẤP 2 (BDD)" in df.columns else ("Họ tên sale", "count"),
        so_hd=("Số hợp đồng", "nunique"),
        doanh_thu=("Doanh thu trước thuế", "sum"),
        est_bonus=("EST_Bonus", "sum") if "EST_Bonus" in df.columns else ("Họ tên sale", "count"),
        affina=("Affina_Revenue", "sum") if "Affina_Revenue" in df.columns else ("Họ tên sale", "count"),
    ).sort_values("doanh_thu", ascending=False)
    result.insert(0, "Rank", range(1, len(result) + 1))
    result = result.rename(columns={
        "chuc_danh": "Chức danh", "source": "Source", "channel": "Channel",
        "bdm": "BDM", "bdd": "BDD",
        "so_hd": "Số HĐ", "doanh_thu": "Doanh thu trước thuế",
        "est_bonus": "EST_Bonus", "affina": "Affina_Revenue",
    })
    return result


def build_by_month(df: pd.DataFrame) -> pd.DataFrame:
    if "Ngày thanh toán" not in df.columns:
        return pd.DataFrame()
    df_copy = df.copy()
    df_copy["Ngày thanh toán"] = pd.to_datetime(df_copy["Ngày thanh toán"], errors="coerce")
    df_copy = df_copy.dropna(subset=["Ngày thanh toán"])
    df_copy["Tháng"] = df_copy["Ngày thanh toán"].dt.to_period("M").astype(str)
    result = df_copy.groupby("Tháng", as_index=False).agg(
        so_hd=("Số hợp đồng", "nunique"),
        so_sale=("Họ tên sale", "nunique"),
        doanh_thu=("Doanh thu trước thuế", "sum"),
        affina=("Affina_Revenue", "sum"),
        bonus=("EST_Bonus", "sum"),
    ).sort_values("Tháng").rename(columns={
        "so_hd": "Số HĐ", "so_sale": "Số Sale",
        "doanh_thu": "Doanh thu trước thuế",
        "affina": "Affina_Revenue", "bonus": "EST_Bonus",
    })
    # Add MoM growth
    if len(result) > 1:
        result["MoM %"] = result["Doanh thu trước thuế"].pct_change() * 100
        result["MoM %"] = result["MoM %"].round(1)
    return result


def build_metadata() -> pd.DataFrame:
    return pd.DataFrame([
        ("Snapshot time (VN)",    datetime.now(VN_TZ).strftime("%Y-%m-%d %H:%M:%S")),
        ("Timezone",              "UTC+7 (Asia/Ho_Chi_Minh)"),
        ("Generated by",          "scripts/export_full_excel.py"),
        ("Data source",           "Supabase - dashboard_master_data"),
        ("Currency",              "VNĐ (Vietnamese Dong)"),
        ("File format",           "Excel .xlsx with openpyxl formatting"),
    ], columns=["Info", "Value"])


# ============================================================================
# MAIN
# ============================================================================
def main():
    log("=" * 70)
    log("AFFINA — EXPORT FULL EXCEL")
    log("=" * 70)

    db_uri = os.environ.get("SUPABASE_DB_URI")
    if not db_uri:
        log("❌ Chưa set SUPABASE_DB_URI trong env hoặc .env")
        sys.exit(1)

    # 1. Load data
    log("1. Kết nối Supabase và tải dashboard_master_data...")
    engine = create_engine(db_uri, pool_pre_ping=True)
    df = pd.read_sql('SELECT * FROM dashboard_master_data', engine)
    log(f"   ✅ Loaded {len(df):,} rows × {len(df.columns)} columns")

    # Convert date columns
    for col in ["Ngày thanh toán", "Ngày bắt đầu", "Ngày kết thúc",
                "Ngày sinh NĐBH", "Ngày sinh NMBH"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    # 2. Build all sheets
    log("2. Đang build 10 sheet phân tích...")
    sheets = {
        "📊 Summary":      build_summary_sheet(df),
        "📁 All_Data":     df,
        "🎨 By_Source":    build_by_source(df),
        "📡 By_Channel":   build_by_channel(df),
        "🛡 By_Loai_BH":  build_by_loai_bh(df),
        "🏢 By_Nha_BH":   build_by_nha_bh(df),
        "📦 Top_Products": build_top_products(df, top_n=30),
        "👥 Sales_Ranking": build_sales_ranking(df),
        "📅 By_Month":     build_by_month(df),
        "ℹ Metadata":     build_metadata(),
    }

    for name, sheet_df in sheets.items():
        log(f"   ✅ {name}: {len(sheet_df):,} rows")

    # 3. Write Excel
    ts = datetime.now(VN_TZ).strftime("%Y%m%d_%H%M")
    filename = f"affina_full_export_{ts}.xlsx"
    log(f"3. Ghi file Excel: {filename}...")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            if sheet_df.empty:
                continue
            # Excel sheet name giới hạn 31 chars, không / \ * ? [ ]
            safe_name = sheet_name
            for ch in ["/", "\\", "?", "*", "[", "]", ":"]:
                safe_name = safe_name.replace(ch, "_")
            safe_name = safe_name[:31]

            # Strip timezone khỏi datetime columns (Excel không hỗ trợ tz-aware)
            sheet_df = _strip_tz(sheet_df)

            sheet_df.to_excel(writer, sheet_name=safe_name, index=False)
            _apply_format(writer.sheets[safe_name], sheet_df,
                          header_color="1F4E78" if sheet_name == "📊 Summary" else "305496")

    # 4. Report
    file_size_mb = os.path.getsize(filename) / (1024 * 1024)
    log("=" * 70)
    log(f"🎉 XONG! File: {filename}")
    log(f"   Kích thước: {file_size_mb:.2f} MB")
    log(f"   Số sheet:   {len([s for s in sheets.values() if not s.empty])}")
    log(f"   Full path:  {os.path.abspath(filename)}")
    log("=" * 70)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"❌ Fatal: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
