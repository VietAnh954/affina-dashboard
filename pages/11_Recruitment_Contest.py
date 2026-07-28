"""
================================================================================
 TRANG 11 — CHƯƠNG TRÌNH THI ĐUA TUYỂN DỤNG "THÊM BẠN, THÊM VUI"
================================================================================
Thời gian: Tháng 06/2026
Đối tượng: CVKD, Trưởng Phòng, Giám Đốc (Core, Neo, TSA)
Giới thiệu ứng viên ngang cấp vào hệ thống.

Mức thưởng:
  - GT Chuyên Viên (DSA/AG/TSA):  100k/người  (DT SP chính ≥ 5tr đến 31/07/2026)
  - GT Trưởng Phòng (BDM/TSA TL): 500k/người  (DT nhóm ≥ 100tr + tuyển ≥ 10 TV đến 31/08)
  - GT Giám Đốc (BDD/TSA Mgr):    1tr/người   (≥ 100% KPI 2 tháng đầu đến 31/08)

Lưu ý: Người GT Chuyên Viên phải có DT > 0 trong T06/2026.
================================================================================
"""
from datetime import date
from io import BytesIO

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from lib.auth import require_auth, render_user_info
from lib.data import (
    COLORS, DATE_COL,
    apply_plotly_layout, empty_state,
    fmt_num, fmt_vnd,
    load_master_data,
)

st.set_page_config(page_title="Thi Đua Tuyển Dụng", layout="wide")

require_auth("recruitment", "Thi Đua Tuyển Dụng — Thêm Bạn Thêm Vui")
render_user_info()

from lib.theme import inject_css, render_header
inject_css()
render_header()


# ============================================================================
# CONFIG
# ============================================================================
CONTEST_NAME = "THÊM BẠN, THÊM VUI"
RECRUIT_MONTH = pd.Timestamp("2026-06-01")
RECRUIT_MONTH_END = pd.Timestamp("2026-06-30")
DEADLINE_CV = pd.Timestamp("2026-07-31")
DEADLINE_TP_GD = pd.Timestamp("2026-08-31")
PAY_DATE_CV = pd.Timestamp("2026-08-10")
PAY_DATE_TP_GD = pd.Timestamp("2026-09-10")

REWARD_CV = 100_000
REWARD_TP = 500_000
REWARD_GD = 1_000_000

CONDITION_CV_REVENUE = 5_000_000
CONDITION_TP_TEAM_REVENUE = 100_000_000
CONDITION_TP_TEAM_SIZE = 10

LEVEL_LABELS = {
    "Chuyen Vien": "Chuyên Viên KD",
    "Truong Phong": "Trưởng Phòng PTKD",
    "Giam Doc": "Giám Đốc PTKD",
}

# "Cấp bậc" trong DSNS → cấp thi đua (dùng cho ngang cấp)
CAP_BAC_MAP = {
    "Nhân viên": "Chuyen Vien",
    "Teamlead": "Truong Phong",
    "Manager": "Giam Doc",
    "Manager_Head": "Giam Doc",
}

# Fallback: "Chức danh" → cấp thi đua (nếu cấp bậc trống)
CHUC_DANH_MAP = {
    "DSA": "Chuyen Vien", "AG": "Chuyen Vien", "TSA": "Chuyen Vien",
    "CTV": "Chuyen Vien", "CVKD": "Chuyen Vien", "RMC": "Chuyen Vien",
    "CTV TSA": "Chuyen Vien",
    "BDM": "Truong Phong", "BDM_DV": "Truong Phong",
    "SM": "Truong Phong", "RMM": "Truong Phong", "UM": "Truong Phong",
    "TSA Team Leader": "Truong Phong", "TSA Leader": "Truong Phong",
    "Teamlead, Agency Managment": "Truong Phong",
    "BDD": "Giam Doc", "BDD_DV": "Giam Doc",
    "SD": "Giam Doc", "RMD": "Giam Doc", "AM": "Giam Doc",
    "TSA Manager": "Giam Doc",
}


# ============================================================================
# HELPERS
# ============================================================================
def _classify_recruit_level(row) -> str:
    """Xác định cấp thi đua: ưu tiên 'Cấp bậc', fallback 'Chức danh'."""
    cap_bac = row.get("Cấp bậc", None) if isinstance(row, dict) else (row["Cấp bậc"] if "Cấp bậc" in row.index else None)
    if pd.notna(cap_bac):
        cb = str(cap_bac).strip()
        if cb in CAP_BAC_MAP:
            return CAP_BAC_MAP[cb]

    chuc_danh = row.get("Chức danh", None) if isinstance(row, dict) else (row["Chức danh"] if "Chức danh" in row.index else None)
    if pd.notna(chuc_danh):
        cd = str(chuc_danh).strip()
        if cd in CHUC_DANH_MAP:
            return CHUC_DANH_MAP[cd]
        cd_upper = cd.upper()
        if "BDD" in cd_upper or "GIAM DOC" in cd_upper:
            return "Giam Doc"
        if "BDM" in cd_upper or "TRUONG" in cd_upper:
            return "Truong Phong"
    return "Chuyen Vien"


def _normalize_phone(val) -> str:
    if pd.isna(val) or str(val).strip() in ("", "nan", "None"):
        return ""
    return str(val).strip().lstrip("0")


def _load_dsns() -> pd.DataFrame | None:
    """Load DSNS trimmed from Supabase for recruitment tracking."""
    try:
        from sqlalchemy import create_engine, text
        import os
        uri = None
        try:
            uri = st.secrets["SUPABASE_DB_URI"]
        except Exception:
            uri = os.environ.get("SUPABASE_DB_URI")
        if not uri:
            return None
        engine = create_engine(uri, pool_pre_ping=True)
        df = pd.read_sql(text("SELECT * FROM dashboard.ds_nhan_su_trim"), engine)
        return df
    except Exception:
        return None


def _export_excel(df_export: pd.DataFrame, sheet_name: str = "Recruitment") -> BytesIO:
    output = BytesIO()
    try:
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        df_export.to_excel(output, index=False, engine="openpyxl")
        output.seek(0)
        return output

    for col in df_export.select_dtypes(include=["datetime64[ns, UTC]", "datetime64[ns]"]).columns:
        df_export[col] = df_export[col].dt.strftime("%d/%m/%Y")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill("solid", fgColor="7038A0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side("thin"), right=Side("thin"),
            top=Side("thin"), bottom=Side("thin")
        )

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            max_len = len(str(cell.value)) + 2
            for row_idx in range(2, min(ws.max_row + 1, 200)):
                data_cell = ws.cell(row=row_idx, column=col_idx)
                data_cell.border = thin_border
                val_len = len(str(data_cell.value)) if data_cell.value else 0
                max_len = max(max_len, val_len + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 35)

    output.seek(0)
    return output


# ============================================================================
# MAIN
# ============================================================================
st.title(f"Thi Dua Tuyen Dung — {CONTEST_NAME}")
st.markdown(
    "**Thoi gian:** Nguoi duoc gioi thieu gia nhap thang 06/2026  |  "
    "**Doi tuong:** CVKD, Truong Phong, Giam Doc (Core, Neo, TSA)"
)

# ============================================================================
# THE LE (expander)
# ============================================================================
with st.expander("**THE LE CHUONG TRINH THI DUA TUYEN DUNG**", expanded=False):
    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown("#### Muc thuong")
        st.markdown("""
| Nguoi duoc gioi thieu | Muc thuong | Dieu kien |
|---|---|---|
| **Chuyen Vien KD** (DSA/AG/TSA) | **100,000 VND**/nguoi | DT SP chinh >= 5 trieu (den 31/07/2026) |
| **Truong Phong PTKD** (BDM/TSA Leader) | **500,000 VND**/nguoi | DT nhom >= 100 trieu + Tuyen >= 10 TV (den 31/08/2026) |
| **Giam Doc PTKD** (BDD/TSA Manager) | **1,000,000 VND**/nguoi | Hoan thanh >= 100% KPI 2 thang dau (den 31/08/2026) |
""")

    with col_right:
        st.markdown("#### Luu y quan trong")
        st.markdown("""
- **Ngang cap**: Chi tinh khi gioi thieu nguoi cung cap bac (CV gioi thieu CV, TP gioi thieu TP, GD gioi thieu GD)
- **Khong ap dung** truong hop cap tren tuyen cap duoi
- Nguoi GT Chuyen Vien phai co **DT > 0 trong T06/2026**
- Nguoi gioi thieu = "Nguoi tuyen dung" trong Affina Pro
- Nguoi duoc GT phai gia nhap **trong thang 06/2026**
""")

    st.markdown("""
**Thoi diem tra thuong:**
- GT Chuyen Vien: truoc 10/08/2026
- GT Truong Phong & Giam Doc: truoc 10/09/2026
""")

st.divider()


# ============================================================================
# LOAD DATA
# ============================================================================
df_all = load_master_data()
if df_all.empty:
    st.warning("Chua co du lieu.")
    st.stop()

df_all[DATE_COL] = pd.to_datetime(df_all[DATE_COL], errors="coerce")

sale_col = "Họ tên sale" if "Họ tên sale" in df_all.columns else "Họ tên"
phone_col = "SĐT sale"

# Load DSNS for recruitment info
dsns = _load_dsns()
has_dsns = dsns is not None and not dsns.empty


# ============================================================================
# TIEN DO TIMELINE
# ============================================================================
st.markdown("### Tien do chuong trinh")

today = pd.Timestamp.now().normalize()

col_t1, col_t2, col_t3, col_t4 = st.columns(4)

with col_t1:
    if today <= RECRUIT_MONTH_END:
        days_left_recruit = (RECRUIT_MONTH_END - today).days
        st.metric("Gia nhap (T06)", f"Con {days_left_recruit} ngay")
    else:
        st.metric("Gia nhap (T06)", "DA KET THUC")

with col_t2:
    if today <= DEADLINE_CV:
        days_left_cv = (DEADLINE_CV - today).days
        st.metric("Deadline CV (31/07)", f"Con {days_left_cv} ngay")
    else:
        st.metric("Deadline CV (31/07)", "DA CHOT")

with col_t3:
    if today <= DEADLINE_TP_GD:
        days_left_tp = (DEADLINE_TP_GD - today).days
        st.metric("Deadline TP/GD (31/08)", f"Con {days_left_tp} ngay")
    else:
        st.metric("Deadline TP/GD (31/08)", "DA CHOT")

with col_t4:
    if today < RECRUIT_MONTH:
        st.metric("Trang thai", "CHUA BAT DAU")
    elif today <= DEADLINE_TP_GD:
        st.metric("Trang thai", "DANG DIEN RA")
    else:
        st.metric("Trang thai", "KET THUC")

# Progress bar
if today < RECRUIT_MONTH:
    pct = 0
elif today > DEADLINE_TP_GD:
    pct = 1.0
else:
    total_days = (DEADLINE_TP_GD - RECRUIT_MONTH).days
    elapsed = (today - RECRUIT_MONTH).days
    pct = min(elapsed / total_days, 1.0)
st.progress(pct)

st.divider()


# ============================================================================
# DATA PROCESSING
# ============================================================================
if has_dsns:
    # ── Process DSNS: normalize phones + classify levels ──
    dsns["_phone"] = dsns["Điện thoại"].apply(_normalize_phone) if "Điện thoại" in dsns.columns else ""
    dsns["_recruiter_phone"] = dsns["Người giới thiệu"].apply(_normalize_phone) if "Người giới thiệu" in dsns.columns else ""

    if "Thời gian bắt đầu" in dsns.columns:
        dsns["join_date"] = pd.to_datetime(dsns["Thời gian bắt đầu"], errors="coerce", dayfirst=True)
    else:
        dsns["join_date"] = pd.NaT

    dsns["level"] = dsns.apply(_classify_recruit_level, axis=1)

    # Channel: prefer converted "Channel", fallback to "Channal"
    if "Channel" not in dsns.columns and "Channal" in dsns.columns:
        dsns["Channel"] = dsns["Channal"]

    # ── New joiners in June 2026 ──
    new_joiners = dsns[
        (dsns["join_date"] >= RECRUIT_MONTH) &
        (dsns["join_date"] <= RECRUIT_MONTH_END)
    ].copy()

    # ── Map recruiter: Code người giới thiệu → Code (primary), phone fallback ──
    has_code_match = "Code người giới thiệu" in new_joiners.columns and "Code" in dsns.columns
    if has_code_match:
        code_to_name = dsns.drop_duplicates(subset=["Code"]).set_index("Code")["Họ tên"].to_dict()
        code_to_level = dsns.drop_duplicates(subset=["Code"]).set_index("Code")["level"].to_dict()
        code_to_phone = dsns.drop_duplicates(subset=["Code"]).set_index("Code")["_phone"].to_dict()
        new_joiners["recruiter_name"] = new_joiners["Code người giới thiệu"].map(code_to_name)
        new_joiners["recruiter_level"] = new_joiners["Code người giới thiệu"].map(code_to_level)
        new_joiners["_recruiter_phone_resolved"] = new_joiners["Code người giới thiệu"].map(code_to_phone)
    else:
        phone_to_name = dsns.drop_duplicates(subset=["_phone"]).set_index("_phone")["Họ tên"].to_dict()
        phone_to_name.pop("", None)
        phone_to_level = dsns.drop_duplicates(subset=["_phone"]).set_index("_phone")["level"].to_dict()
        phone_to_level.pop("", None)
        new_joiners["recruiter_name"] = new_joiners["_recruiter_phone"].map(phone_to_name)
        new_joiners["recruiter_level"] = new_joiners["_recruiter_phone"].map(phone_to_level)
        new_joiners["_recruiter_phone_resolved"] = new_joiners["_recruiter_phone"]

    # Ngang cấp check
    new_joiners["is_same_level"] = new_joiners["level"] == new_joiners["recruiter_level"]

    # ── Revenue data (from master_data, match by phone) ──
    df_revenue = df_all[df_all[DATE_COL].notna()].copy()
    df_revenue["_sale_phone"] = df_revenue[phone_col].apply(_normalize_phone) if phone_col in df_revenue.columns else ""

    # Recruiter revenue in June 2026 (condition for CV level referrers)
    june_revenue = df_revenue[
        (df_revenue[DATE_COL] >= RECRUIT_MONTH) &
        (df_revenue[DATE_COL] <= RECRUIT_MONTH_END)
    ].groupby("_sale_phone")["Doanh thu trước thuế"].sum().reset_index()
    june_revenue.columns = ["_phone", "recruiter_june_revenue"]

    # New joiner revenue (from join to CV deadline 31/07)
    joiner_cv_revenue = df_revenue[
        (df_revenue[DATE_COL] >= RECRUIT_MONTH) &
        (df_revenue[DATE_COL] <= DEADLINE_CV)
    ].groupby("_sale_phone")["Doanh thu trước thuế"].sum().reset_index()
    joiner_cv_revenue.columns = ["_phone", "joiner_revenue_to_jul"]

    # New joiner revenue (from join to TP/GD deadline 31/08)
    joiner_tp_gd_revenue = df_revenue[
        (df_revenue[DATE_COL] >= RECRUIT_MONTH) &
        (df_revenue[DATE_COL] <= DEADLINE_TP_GD)
    ].groupby("_sale_phone")["Doanh thu trước thuế"].sum().reset_index()
    joiner_tp_gd_revenue.columns = ["_phone", "joiner_revenue_to_aug"]

    # Merge revenue into new_joiners
    new_joiners = new_joiners.merge(joiner_cv_revenue, on="_phone", how="left")
    new_joiners = new_joiners.merge(joiner_tp_gd_revenue, on="_phone", how="left")
    new_joiners["joiner_revenue_to_jul"] = new_joiners["joiner_revenue_to_jul"].fillna(0)
    new_joiners["joiner_revenue_to_aug"] = new_joiners["joiner_revenue_to_aug"].fillna(0)

    # Recruiter june revenue (match by resolved phone)
    rec_phone_col = "_recruiter_phone_resolved" if "_recruiter_phone_resolved" in new_joiners.columns else "_recruiter_phone"
    new_joiners = new_joiners.merge(
        june_revenue, left_on=rec_phone_col, right_on="_phone",
        how="left", suffixes=("", "_rec")
    )
    new_joiners["recruiter_june_revenue"] = new_joiners["recruiter_june_revenue"].fillna(0)

    # ── Eligibility check ──
    def _check_eligibility(row):
        level = row["level"]
        is_same = row["is_same_level"]
        if not is_same:
            return "Khong du DK (khac cap)"

        if level == "Chuyen Vien":
            if row["recruiter_june_revenue"] <= 0:
                return "Nguoi GT chua co DT T06"
            if row["joiner_revenue_to_jul"] >= CONDITION_CV_REVENUE:
                return "DAT"
            else:
                return f"Chua dat (DT: {fmt_vnd(row['joiner_revenue_to_jul'], short=True)}/{fmt_vnd(CONDITION_CV_REVENUE, short=True)})"

        elif level == "Truong Phong":
            if row["joiner_revenue_to_aug"] >= CONDITION_TP_TEAM_REVENUE:
                return "Dat DT (can kiem tra so TV)"
            else:
                return f"Chua dat DT ({fmt_vnd(row['joiner_revenue_to_aug'], short=True)}/{fmt_vnd(CONDITION_TP_TEAM_REVENUE, short=True)})"

        elif level == "Giam Doc":
            return "Can kiem tra KPI 2 thang"

        return "N/A"

    new_joiners["trang_thai"] = new_joiners.apply(_check_eligibility, axis=1)

    # ============================================================================
    # DISPLAY — NEW JOINERS
    # ============================================================================
    st.markdown("### Nguoi duoc gioi thieu — Gia nhap T06/2026")

    n_total = len(new_joiners)
    n_same_level = new_joiners["is_same_level"].sum()
    n_eligible = (new_joiners["trang_thai"] == "DAT").sum()

    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    col_m1.metric("Tong nguoi moi T06", fmt_num(n_total))
    col_m2.metric("Ngang cap (hop le)", fmt_num(n_same_level))
    col_m3.metric("Da dat dieu kien", fmt_num(n_eligible))
    col_m4.metric("Tong thuong du kien", fmt_vnd(n_eligible * REWARD_CV, short=True))

    # Filter by level
    filter_level = st.selectbox(
        "Loc theo cap bac nguoi duoc GT",
        options=["Tat ca", "Chuyen Vien", "Truong Phong", "Giam Doc"],
        index=0, key="recruit_level_filter"
    )

    display_joiners = new_joiners.copy()
    if filter_level != "Tat ca":
        display_joiners = display_joiners[display_joiners["level"] == filter_level]

    if not display_joiners.empty:
        disp_cols = {
            "Họ tên": "Nguoi duoc GT",
            "Cấp bậc": "Cap bac (DSNS)",
            "Chức danh": "Chuc danh",
            "level": "Cap thi dua",
            "Channel": "Kenh",
            "Phòng ban": "Phong ban",
            "join_date": "Ngay gia nhap",
            "recruiter_name": "Nguoi gioi thieu",
            "recruiter_level": "Cap nguoi GT",
            "is_same_level": "Ngang cap",
            "joiner_revenue_to_jul": "DT den 31/07",
            "joiner_revenue_to_aug": "DT den 31/08",
            "trang_thai": "Trang thai",
        }
        disp_df = display_joiners[[c for c in disp_cols.keys() if c in display_joiners.columns]].copy()
        disp_df = disp_df.rename(columns=disp_cols)

        if "Ngay gia nhap" in disp_df.columns:
            disp_df["Ngay gia nhap"] = disp_df["Ngay gia nhap"].dt.strftime("%d/%m/%Y")
        if "DT den 31/07" in disp_df.columns:
            disp_df["DT den 31/07"] = disp_df["DT den 31/07"].apply(lambda v: fmt_vnd(v, short=True))
        if "DT den 31/08" in disp_df.columns:
            disp_df["DT den 31/08"] = disp_df["DT den 31/08"].apply(lambda v: fmt_vnd(v, short=True))
        if "Ngang cap" in disp_df.columns:
            disp_df["Ngang cap"] = disp_df["Ngang cap"].map({True: "Co", False: "Khong"})

        st.dataframe(disp_df, hide_index=True, use_container_width=True, height=400)
    else:
        empty_state("Khong co nguoi moi gia nhap trong T06/2026 (hoac chua co du lieu DSNS).")

    st.divider()

    # ============================================================================
    # RECRUITER LEADERBOARD
    # ============================================================================
    st.markdown("### Bang xep hang Nguoi gioi thieu")

    # Only count same-level referrals
    valid_referrals = new_joiners[new_joiners["is_same_level"]].copy()

    if not valid_referrals.empty:
        recruiter_board = valid_referrals.groupby(["recruiter_name", "recruiter_level"]).agg(
            so_nguoi_gt=("Họ tên", "count"),
            so_dat_dk=("trang_thai", lambda x: (x == "DAT").sum()),
            tong_dt_nguoi_moi=("joiner_revenue_to_jul", "sum"),
        ).reset_index()

        # Calculate reward
        def _calc_reward(row):
            level = row["recruiter_level"]
            n = int(row["so_dat_dk"])
            if level == "Chuyen Vien":
                return n * REWARD_CV
            elif level == "Truong Phong":
                return n * REWARD_TP
            elif level == "Giam Doc":
                return n * REWARD_GD
            return 0

        recruiter_board["thuong_du_kien"] = recruiter_board.apply(_calc_reward, axis=1)
        recruiter_board = recruiter_board.sort_values(
            ["so_dat_dk", "so_nguoi_gt"], ascending=[False, False]
        ).reset_index(drop=True)
        recruiter_board.insert(0, "Hang", range(1, len(recruiter_board) + 1))

        # Display
        disp_board = recruiter_board.rename(columns={
            "recruiter_name": "Nguoi gioi thieu",
            "recruiter_level": "Cap bac",
            "so_nguoi_gt": "So nguoi GT",
            "so_dat_dk": "Dat DK",
            "tong_dt_nguoi_moi": "Tong DT nguoi moi",
            "thuong_du_kien": "Thuong du kien",
        })
        disp_board["Tong DT nguoi moi"] = disp_board["Tong DT nguoi moi"].apply(lambda v: fmt_vnd(v, short=True))
        disp_board["Thuong du kien"] = disp_board["Thuong du kien"].apply(lambda v: fmt_vnd(v))
        disp_board["Cap bac"] = disp_board["Cap bac"].map(LEVEL_LABELS).fillna(disp_board["Cap bac"])

        st.dataframe(disp_board, hide_index=True, use_container_width=True)

        # Top 3 highlight
        top3 = recruiter_board.head(3)
        if not top3.empty:
            cols_top = st.columns(min(3, len(top3)))
            medals = ["1.", "2.", "3."]
            for i, (_, row) in enumerate(top3.iterrows()):
                if i < len(cols_top):
                    with cols_top[i]:
                        st.metric(
                            f"{medals[i]} {row['recruiter_name']}",
                            f"{int(row['so_dat_dk'])} dat / {int(row['so_nguoi_gt'])} GT",
                            delta=f"Thuong: {fmt_vnd(row['thuong_du_kien'], short=True)}",
                        )
    else:
        empty_state("Chua co du lieu gioi thieu ngang cap.")

    st.divider()

    # ============================================================================
    # CHARTS — TONG QUAN
    # ============================================================================
    st.markdown("### Thong ke tong quan tuyen dung")

    # ── Chart 1 & 2: Pie + Bar overview (with shared filter) ──
    ch_ov_f1, ch_ov_f2 = st.columns(2)
    with ch_ov_f1:
        ov_channel_opts = ["Tat ca"] + sorted(new_joiners["Channel"].dropna().unique().tolist()) if "Channel" in new_joiners.columns else ["Tat ca"]
        ov_ch_filter = st.selectbox("Kenh (tong quan)", options=ov_channel_opts, index=0, key="ov_ch_f")
    with ch_ov_f2:
        ov_same_filter = st.selectbox("Chi ngang cap?", options=["Tat ca", "Chi ngang cap", "Chi khac cap"], index=0, key="ov_same_f")

    ov_data = new_joiners.copy()
    if ov_ch_filter != "Tat ca" and "Channel" in ov_data.columns:
        ov_data = ov_data[ov_data["Channel"] == ov_ch_filter]
    if ov_same_filter == "Chi ngang cap":
        ov_data = ov_data[ov_data["is_same_level"]]
    elif ov_same_filter == "Chi khac cap":
        ov_data = ov_data[~ov_data["is_same_level"]]

    col_chart1, col_chart2 = st.columns(2)
    with col_chart1:
        if not ov_data.empty:
            by_level = ov_data.groupby("level").size().reset_index(name="count")
            by_level["level_label"] = by_level["level"].map(LEVEL_LABELS)
            fig = px.pie(
                by_level, names="level_label", values="count",
                color_discrete_sequence=["#B44BC8", "#F06EC2", "#8B6FC9"],
            )
            fig.update_traces(textposition="inside", textinfo="value+percent")
            st.plotly_chart(apply_plotly_layout(fig, title="Nguoi moi theo cap bac", height=350), use_container_width=True)
        else:
            empty_state("Khong co du lieu.")

    with col_chart2:
        if not ov_data.empty and "Channel" in ov_data.columns:
            by_channel = ov_data.groupby("Channel").size().reset_index(name="count")
            fig2 = px.bar(
                by_channel, x="Channel", y="count", color="Channel",
                color_discrete_map={"Core": COLORS["Core"], "Neo": COLORS["Neo"], "TSA": COLORS["TSA"]},
            )
            fig2.update_layout(showlegend=False)
            st.plotly_chart(apply_plotly_layout(fig2, title="Tuyen dung theo kenh", height=350), use_container_width=True)
        else:
            empty_state("Khong co du lieu Channel.")

    st.divider()

    # ── Chart 3: Funnel — Tong tuyen → Ngang cap → Dat DT → Dat DK ──
    st.markdown("### Funnel tuyen dung")
    fn_f1, fn_f2 = st.columns(2)
    with fn_f1:
        fn_level = st.selectbox("Cap bac (funnel)", options=["Tat ca", "Chuyen Vien", "Truong Phong", "Giam Doc"], index=0, key="fn_lv_f")
    with fn_f2:
        fn_ch = st.selectbox("Kenh (funnel)", options=ov_channel_opts, index=0, key="fn_ch_f")

    fn_data = new_joiners.copy()
    if fn_level != "Tat ca":
        fn_data = fn_data[fn_data["level"] == fn_level]
    if fn_ch != "Tat ca" and "Channel" in fn_data.columns:
        fn_data = fn_data[fn_data["Channel"] == fn_ch]

    if not fn_data.empty:
        funnel_stages = [
            ("Tong nguoi moi T06", len(fn_data)),
            ("Co nguoi gioi thieu", fn_data["recruiter_name"].notna().sum()),
            ("Ngang cap", fn_data["is_same_level"].sum()),
            ("Dat dieu kien", (fn_data["trang_thai"] == "DAT").sum()),
        ]
        fig_fn = go.Figure(go.Funnel(
            y=[s[0] for s in funnel_stages],
            x=[s[1] for s in funnel_stages],
            textinfo="value+percent initial",
            marker=dict(color=["#8B6FC9", "#B44BC8", "#F06EC2", "#5FBFA0"]),
        ))
        st.plotly_chart(apply_plotly_layout(fig_fn, title="", height=350), use_container_width=True)
    else:
        empty_state("Khong co du lieu.")

    st.divider()

    # ── Chart 4: Tien do DT Chuyen Vien moi (target 5 trieu) ──
    st.markdown("### Tien do DT Chuyen Vien moi")
    cv_f1, cv_f2, cv_f3 = st.columns(3)
    with cv_f1:
        cv_ch = st.selectbox("Kenh (CV)", options=ov_channel_opts, index=0, key="cv_ch_f")
    with cv_f2:
        cv_status = st.selectbox("Trang thai (CV)", options=["Tat ca", "Dat", "Chua dat"], index=0, key="cv_st_f")
    with cv_f3:
        cv_n_show = st.slider("So nguoi hien thi", min_value=5, max_value=50, value=20, key="cv_n_show")

    cv_joiners = new_joiners[(new_joiners["level"] == "Chuyen Vien") & (new_joiners["is_same_level"])].copy()
    if cv_ch != "Tat ca" and "Channel" in cv_joiners.columns:
        cv_joiners = cv_joiners[cv_joiners["Channel"] == cv_ch]
    cv_joiners["pct_target"] = (cv_joiners["joiner_revenue_to_jul"] / CONDITION_CV_REVENUE * 100).clip(0, 200)
    if cv_status == "Dat":
        cv_joiners = cv_joiners[cv_joiners["pct_target"] >= 100]
    elif cv_status == "Chua dat":
        cv_joiners = cv_joiners[cv_joiners["pct_target"] < 100]

    if not cv_joiners.empty:
        cv_chart = cv_joiners[["Họ tên", "pct_target", "joiner_revenue_to_jul"]].sort_values(
            "pct_target", ascending=True
        ).tail(cv_n_show)

        fig3 = go.Figure()
        fig3.add_trace(go.Bar(
            y=cv_chart["Họ tên"], x=cv_chart["pct_target"],
            orientation="h",
            marker_color=cv_chart["pct_target"].apply(
                lambda p: "#5FBFA0" if p >= 100 else ("#EDB16E" if p >= 50 else "#E8738F")
            ),
            text=cv_chart["joiner_revenue_to_jul"].apply(lambda v: fmt_vnd(v, short=True)),
            textposition="outside",
        ))
        fig3.add_vline(x=100, line_dash="dash", line_color="#B44BC8", annotation_text="Target 5tr")
        fig3.update_layout(xaxis_title="% Target", yaxis_title="", height=max(300, 30 * len(cv_chart)))
        st.plotly_chart(apply_plotly_layout(fig3, title=""), use_container_width=True)

        st.caption(f"Hien thi {len(cv_chart)}/{len(cv_joiners)} nguoi | Xanh = Dat, Vang = >=50%, Hong = <50%")
    else:
        empty_state("Khong co Chuyen Vien moi trong bo loc nay.")

    st.divider()

    # ── Chart 5: Phan bo DT nguoi moi (Histogram) ──
    st.markdown("### Phan bo doanh thu nguoi moi")
    hs_f1, hs_f2 = st.columns(2)
    with hs_f1:
        hs_level = st.selectbox("Cap bac (phan bo)", options=["Tat ca", "Chuyen Vien", "Truong Phong", "Giam Doc"], index=0, key="hs_lv_f")
    with hs_f2:
        hs_ch = st.selectbox("Kenh (phan bo)", options=ov_channel_opts, index=0, key="hs_ch_f")

    hs_data = new_joiners[new_joiners["is_same_level"]].copy()
    if hs_level != "Tat ca":
        hs_data = hs_data[hs_data["level"] == hs_level]
    if hs_ch != "Tat ca" and "Channel" in hs_data.columns:
        hs_data = hs_data[hs_data["Channel"] == hs_ch]

    if not hs_data.empty and hs_data["joiner_revenue_to_jul"].sum() > 0:
        fig_hs = px.histogram(
            hs_data, x="joiner_revenue_to_jul", nbins=20,
            color_discrete_sequence=["#B44BC8"],
            labels={"joiner_revenue_to_jul": "Doanh thu (VND)"},
        )
        fig_hs.add_vline(x=CONDITION_CV_REVENUE, line_dash="dash", line_color="#5FBFA0", annotation_text="5tr")
        st.plotly_chart(apply_plotly_layout(fig_hs, title="", height=350), use_container_width=True)
        st.caption(f"Trung binh: {fmt_vnd(hs_data['joiner_revenue_to_jul'].mean(), short=True)} | "
                   f"Trung vi: {fmt_vnd(hs_data['joiner_revenue_to_jul'].median(), short=True)}")
    else:
        empty_state("Khong co du lieu DT.")

    st.divider()

    # ── Chart 6: Top Recruiters (horizontal bar) ──
    st.markdown("### Top nguoi gioi thieu")
    tr_f1, tr_f2 = st.columns(2)
    with tr_f1:
        tr_level = st.selectbox("Cap bac (recruiter)", options=["Tat ca", "Chuyen Vien", "Truong Phong", "Giam Doc"], index=0, key="tr_lv_f")
    with tr_f2:
        tr_n = st.slider("So nguoi hien thi", min_value=5, max_value=30, value=10, key="tr_n_show")

    tr_data = valid_referrals.copy()
    if tr_level != "Tat ca":
        tr_data = tr_data[tr_data["recruiter_level"] == tr_level]

    if not tr_data.empty:
        tr_board = tr_data.groupby("recruiter_name").agg(
            so_gt=("Họ tên", "count"),
            so_dat=("trang_thai", lambda x: (x == "DAT").sum()),
        ).reset_index().sort_values("so_gt", ascending=True).tail(tr_n)

        fig_tr = go.Figure()
        fig_tr.add_trace(go.Bar(
            y=tr_board["recruiter_name"], x=tr_board["so_dat"],
            name="Dat DK", orientation="h", marker_color="#5FBFA0",
            text=tr_board["so_dat"], textposition="inside",
        ))
        fig_tr.add_trace(go.Bar(
            y=tr_board["recruiter_name"], x=tr_board["so_gt"] - tr_board["so_dat"],
            name="Chua dat", orientation="h", marker_color="#EDB16E",
            text=(tr_board["so_gt"] - tr_board["so_dat"]), textposition="inside",
        ))
        fig_tr.update_layout(barmode="stack", height=max(300, 35 * len(tr_board)))
        st.plotly_chart(apply_plotly_layout(fig_tr, title="", height=max(300, 35 * len(tr_board))), use_container_width=True)
    else:
        empty_state("Khong co du lieu recruiter.")

    st.divider()

    # ── Chart 7: Tien trinh tuyen dung theo tuan trong T06 ──
    st.markdown("### Tien trinh tuyen dung theo tuan T06/2026")
    wk_f1, wk_f2 = st.columns(2)
    with wk_f1:
        wk_level = st.selectbox("Cap bac (tuan)", options=["Tat ca", "Chuyen Vien", "Truong Phong", "Giam Doc"], index=0, key="wk_lv_f")
    with wk_f2:
        wk_ch = st.selectbox("Kenh (tuan)", options=ov_channel_opts, index=0, key="wk_ch_f")

    wk_data = new_joiners.copy()
    if wk_level != "Tat ca":
        wk_data = wk_data[wk_data["level"] == wk_level]
    if wk_ch != "Tat ca" and "Channel" in wk_data.columns:
        wk_data = wk_data[wk_data["Channel"] == wk_ch]

    if not wk_data.empty and wk_data["join_date"].notna().any():
        wk_data["week"] = wk_data["join_date"].dt.isocalendar().week.astype(int)
        wk_data["week_label"] = wk_data["join_date"].apply(
            lambda d: f"Tuan {d.day // 7 + 1} ({d.strftime('%d/%m')})" if pd.notna(d) else ""
        )
        by_week = wk_data.groupby("week").agg(
            count=("Họ tên", "count"),
            week_start=("join_date", "min"),
        ).reset_index().sort_values("week")
        by_week["label"] = by_week["week_start"].dt.strftime("Tuan %d/%m")

        fig_wk = px.bar(
            by_week, x="label", y="count",
            color_discrete_sequence=["#B44BC8"],
            labels={"label": "Tuan", "count": "So nguoi gia nhap"},
            text="count",
        )
        fig_wk.update_traces(textposition="outside")
        st.plotly_chart(apply_plotly_layout(fig_wk, title="", height=350), use_container_width=True)
    else:
        empty_state("Khong co du lieu ngay gia nhap.")

    st.divider()

    # ── Chart 8: Ngan sach thuong theo kenh (stacked bar) ──
    st.markdown("### Ngan sach thuong du kien theo kenh")
    bg_filter = st.selectbox("Cap bac (ngan sach)", options=["Tat ca", "Chuyen Vien", "Truong Phong", "Giam Doc"], index=0, key="bg_lv_f")

    bg_data = new_joiners[new_joiners["is_same_level"] & (new_joiners["trang_thai"] == "DAT")].copy()
    if bg_filter != "Tat ca":
        bg_data = bg_data[bg_data["level"] == bg_filter]

    if not bg_data.empty and "Channel" in bg_data.columns:
        def _reward_for_level(lv):
            return {"Chuyen Vien": REWARD_CV, "Truong Phong": REWARD_TP, "Giam Doc": REWARD_GD}.get(lv, 0)

        bg_data["reward"] = bg_data["level"].apply(_reward_for_level)
        bg_agg = bg_data.groupby(["Channel", "level"]).agg(
            count=("Họ tên", "count"),
            total_reward=("reward", "sum"),
        ).reset_index()
        bg_agg["level_label"] = bg_agg["level"].map(LEVEL_LABELS)

        fig_bg = px.bar(
            bg_agg, x="Channel", y="total_reward", color="level_label",
            barmode="stack", text="count",
            color_discrete_map={
                LEVEL_LABELS["Chuyen Vien"]: "#B44BC8",
                LEVEL_LABELS["Truong Phong"]: "#F06EC2",
                LEVEL_LABELS["Giam Doc"]: "#8B6FC9",
            },
            labels={"total_reward": "Thuong (VND)", "Channel": "Kenh", "level_label": "Cap bac"},
        )
        fig_bg.update_traces(textposition="inside")
        st.plotly_chart(apply_plotly_layout(fig_bg, title="", height=380), use_container_width=True)
    elif not bg_data.empty:
        st.metric("Tong thuong (khong co Channel)", fmt_vnd(bg_data["level"].apply(_reward_for_level).sum()))
    else:
        empty_state("Chua co nguoi dat dieu kien.")

    st.divider()

    # ── Chart 9: Ty le Dat/Chua dat theo cap bac (stacked %) ──
    st.markdown("### Ty le dat dieu kien theo cap bac")
    el_ch = st.selectbox("Kenh (ty le)", options=ov_channel_opts, index=0, key="el_ch_f")

    el_data = new_joiners[new_joiners["is_same_level"]].copy()
    if el_ch != "Tat ca" and "Channel" in el_data.columns:
        el_data = el_data[el_data["Channel"] == el_ch]

    if not el_data.empty:
        el_data["is_dat"] = el_data["trang_thai"] == "DAT"
        el_agg = el_data.groupby("level").agg(
            dat=("is_dat", "sum"),
            total=("is_dat", "count"),
        ).reset_index()
        el_agg["chua_dat"] = el_agg["total"] - el_agg["dat"]
        el_agg["level_label"] = el_agg["level"].map(LEVEL_LABELS)
        el_agg["pct_dat"] = (el_agg["dat"] / el_agg["total"] * 100).round(1)

        fig_el = go.Figure()
        fig_el.add_trace(go.Bar(
            x=el_agg["level_label"], y=el_agg["dat"],
            name="Dat DK", marker_color="#5FBFA0",
            text=el_agg.apply(lambda r: f"{int(r['dat'])} ({r['pct_dat']}%)", axis=1),
            textposition="inside",
        ))
        fig_el.add_trace(go.Bar(
            x=el_agg["level_label"], y=el_agg["chua_dat"],
            name="Chua dat", marker_color="#E8738F",
            text=el_agg["chua_dat"].astype(int), textposition="inside",
        ))
        fig_el.update_layout(barmode="stack")
        st.plotly_chart(apply_plotly_layout(fig_el, title="", height=380), use_container_width=True)
    else:
        empty_state("Khong co du lieu ngang cap.")

    st.divider()

    # ============================================================================
    # SUMMARY & BUDGET
    # ============================================================================
    st.markdown("### Tong ket & Ngan sach du kien")

    summary_data = []
    for level in ["Chuyen Vien", "Truong Phong", "Giam Doc"]:
        level_data = new_joiners[(new_joiners["level"] == level) & (new_joiners["is_same_level"])]
        n_gt = len(level_data)
        n_dat = (level_data["trang_thai"] == "DAT").sum() if not level_data.empty else 0
        reward = {
            "Chuyen Vien": REWARD_CV,
            "Truong Phong": REWARD_TP,
            "Giam Doc": REWARD_GD,
        }[level]
        summary_data.append({
            "Cap bac": LEVEL_LABELS[level],
            "So nguoi GT": n_gt,
            "Dat DK": n_dat,
            "Muc thuong/nguoi": fmt_vnd(reward),
            "Tong thuong": fmt_vnd(n_dat * reward),
        })

    summary_df = pd.DataFrame(summary_data)
    total_budget = sum(
        (new_joiners[(new_joiners["level"] == lv) & (new_joiners["is_same_level"]) & (new_joiners["trang_thai"] == "DAT")].shape[0])
        * rw for lv, rw in [("Chuyen Vien", REWARD_CV), ("Truong Phong", REWARD_TP), ("Giam Doc", REWARD_GD)]
    )

    st.dataframe(summary_df, hide_index=True, use_container_width=True)
    st.metric("Tong ngan sach thuong du kien", fmt_vnd(total_budget))

    st.divider()

    # ============================================================================
    # EXPORT EXCEL
    # ============================================================================
    st.markdown("### Xuat report Excel")

    export_choice = st.radio(
        "Chon bao cao",
        options=["Danh sach nguoi duoc GT", "Bang xep hang nguoi GT", "Tong ket"],
        horizontal=True,
        key="recruit_export_choice",
    )

    if export_choice == "Danh sach nguoi duoc GT":
        exp_base_cols = ["Họ tên", "Cấp bậc", "Chức danh", "level", "Channel", "Phòng ban",
                         "join_date", "recruiter_name", "recruiter_level",
                         "is_same_level", "joiner_revenue_to_jul", "joiner_revenue_to_aug", "trang_thai"]
        exp_cols = [c for c in exp_base_cols if c in new_joiners.columns]
        exp_df = new_joiners[exp_cols].copy()
        exp_rename = {
            "Họ tên": "Nguoi duoc GT", "Cấp bậc": "Cap bac (DSNS)", "Chức danh": "Chuc danh",
            "level": "Cap thi dua", "Channel": "Kenh", "Phòng ban": "Phong ban",
            "join_date": "Ngay gia nhap", "recruiter_name": "Nguoi GT", "recruiter_level": "Cap nguoi GT",
            "is_same_level": "Ngang cap", "joiner_revenue_to_jul": "DT den 31/07",
            "joiner_revenue_to_aug": "DT den 31/08", "trang_thai": "Trang thai",
        }
        exp_df = exp_df.rename(columns=exp_rename)
        exp_df["Ngang cap"] = exp_df["Ngang cap"].map({True: "Co", False: "Khong"})
        sheet = "DS Nguoi duoc GT"
    elif export_choice == "Bang xep hang nguoi GT" and not valid_referrals.empty:
        exp_df = recruiter_board.drop(columns=["Hang"], errors="ignore").rename(columns={
            "recruiter_name": "Nguoi GT", "recruiter_level": "Cap bac",
            "so_nguoi_gt": "So nguoi GT", "so_dat_dk": "Dat DK",
            "tong_dt_nguoi_moi": "Tong DT nguoi moi", "thuong_du_kien": "Thuong du kien",
        })
        sheet = "BXH Nguoi GT"
    else:
        exp_df = summary_df.copy()
        sheet = "Tong ket"

    excel_data = _export_excel(exp_df, sheet_name=sheet)
    st.download_button(
        label=f"Tai Excel — {export_choice}",
        data=excel_data,
        file_name=f"Recruitment_Contest_{sheet.replace(' ', '_')}_{today.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="recruit_download_excel",
    )

else:
    # ── FALLBACK: No DSNS data — use heuristic from master_data ──
    st.warning(
        "Chua co du lieu DSNS (bang `ds_nhan_su_trim` tren Supabase). "
        "Chay lai ETL (`build_dashboard_data.py`) de push du lieu nhan su. "
        "Dang hien thi thong tin co ban tu doanh thu."
    )

    st.markdown("### Thong ke DT thang 06/2026")

    df_june = df_all[
        (df_all[DATE_COL] >= RECRUIT_MONTH) & (df_all[DATE_COL] <= RECRUIT_MONTH_END)
    ].copy()

    if not df_june.empty:
        june_by_sale = df_june.groupby(sale_col).agg(
            revenue=("Doanh thu trước thuế", "sum"),
            n_hd=("Số hợp đồng", "nunique") if "Số hợp đồng" in df_june.columns else (sale_col, "count"),
        ).reset_index().sort_values("revenue", ascending=False)

        col_j1, col_j2, col_j3 = st.columns(3)
        col_j1.metric("So sale co DT T06", fmt_num(len(june_by_sale)))
        col_j2.metric("Tong DT T06", fmt_vnd(june_by_sale["revenue"].sum(), short=True))
        col_j3.metric("TB DT/sale", fmt_vnd(june_by_sale["revenue"].mean(), short=True))

        st.dataframe(
            june_by_sale.head(30).rename(columns={
                sale_col: "Ho ten", "revenue": "DT T06", "n_hd": "So HD"
            }),
            hide_index=True, use_container_width=True
        )
    else:
        empty_state("Chua co du lieu doanh thu thang 06/2026.")

    # First-time sales (heuristic for new joiners)
    st.markdown("### Nguoi moi (du doan) — Lan dau co DT trong T06/2026")
    st.caption("Day la danh sach sale co giao dich dau tien trong T06/2026 (khong co giao dich nao truoc do). Day la uoc tinh, khong chinh xac 100%.")

    first_tx = df_all.groupby(sale_col)[DATE_COL].min().reset_index()
    first_tx.columns = [sale_col, "first_tx_date"]
    new_in_june = first_tx[
        (first_tx["first_tx_date"] >= RECRUIT_MONTH) &
        (first_tx["first_tx_date"] <= RECRUIT_MONTH_END)
    ]

    if not new_in_june.empty:
        new_june_revenue = df_all[
            (df_all[sale_col].isin(new_in_june[sale_col])) &
            (df_all[DATE_COL] >= RECRUIT_MONTH) &
            (df_all[DATE_COL] <= DEADLINE_CV)
        ].groupby(sale_col)["Doanh thu trước thuế"].sum().reset_index()
        new_june_revenue.columns = [sale_col, "revenue_to_jul"]

        new_june_df = new_in_june.merge(new_june_revenue, on=sale_col, how="left")
        new_june_df["revenue_to_jul"] = new_june_df["revenue_to_jul"].fillna(0)
        new_june_df["dat_5tr"] = new_june_df["revenue_to_jul"] >= CONDITION_CV_REVENUE
        new_june_df = new_june_df.sort_values("revenue_to_jul", ascending=False).reset_index(drop=True)

        col_n1, col_n2, col_n3 = st.columns(3)
        col_n1.metric("So nguoi moi (uoc tinh)", fmt_num(len(new_june_df)))
        col_n2.metric("Dat >= 5tr DT", fmt_num(int(new_june_df["dat_5tr"].sum())))
        col_n3.metric("Chua dat", fmt_num(int((~new_june_df["dat_5tr"]).sum())))

        disp_new = new_june_df.rename(columns={
            sale_col: "Ho ten", "first_tx_date": "GD dau tien",
            "revenue_to_jul": "DT den 31/07", "dat_5tr": "Dat 5tr"
        })
        disp_new["GD dau tien"] = pd.to_datetime(disp_new["GD dau tien"]).dt.strftime("%d/%m/%Y")
        disp_new["DT den 31/07"] = disp_new["DT den 31/07"].apply(lambda v: fmt_vnd(v, short=True))
        disp_new["Dat 5tr"] = disp_new["Dat 5tr"].map({True: "Dat", False: "Chua"})
        st.dataframe(disp_new, hide_index=True, use_container_width=True)
    else:
        empty_state("Khong tim thay nguoi moi trong T06/2026.")


st.divider()

# ============================================================================
# FOOTER
# ============================================================================
st.markdown(
    "---\n"
    "*Chuong trinh thi dua tuyen dung T06/2026. Du lieu cap nhat hang ngay. "
    "Ket qua chinh thuc do Phong Nhan su cong bo.*"
)
