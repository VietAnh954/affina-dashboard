"""
================================================================================
 TRANG 10 — KPI COMPETITION: CLB Tinh Hoa Affina 2026-2027
================================================================================
Chương trình thi đua: 01/04/2026 - 31/03/2027
13 suất du lịch Trung Quốc

Cấp bậc:
  - Giám Đốc (BDD): Top 3 KPI QL, >=70% x 3 tháng, TB >=50%
  - Trưởng Phòng (BDM): Top 5 KPI QL, >=70% x 3 tháng, TB >=50%
  - Chuyên Viên (CVKD/AG-RMC/CTV TSA):  Top 5 điểm quy đổi

KPI Quản lý (CORE only):
  - BDM (Trưởng Phòng): target 200,000,000 VNĐ/tháng (Số tiền thanh toán team)
  - BDD (Giám Đốc):     target 450,000,000 VNĐ/tháng (Số tiền thanh toán team)
  - NEO/TSA: chỉ xếp hạng theo DT team, không tính % KPI
  - BDH: không tham gia

Tính điểm Chuyên Viên:
  - Mỗi 5 triệu doanh thu cá nhân = 1 điểm (tính theo tháng)
  - Mỗi 5 triệu doanh thu từ người giới thiệu = 1 điểm (tính theo tháng)
  - Bonus rank THEO QUÝ: Hạng 1: +10, Hạng 2: +5, Hạng 3: +3
  - Tiebreaker: bằng điểm → ai có tổng doanh thu cao hơn xếp trên

LƯU Ý: KHÔNG hiển thị Affina_Revenue trong trang này.
================================================================================
"""
from datetime import date
from io import BytesIO

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from lib.auth import require_auth, render_user_info
from lib.data import (
    COLORS, DATE_COL,
    apply_plotly_layout, empty_state,
    fmt_num, fmt_vnd,
    load_master_data,
)

st.set_page_config(page_title="KPI Competition", layout="wide")

require_auth("kpi", "KPI Competition — CLB Tinh Hoa Affina")
render_user_info()

from lib.theme import inject_css, render_header
inject_css()
render_header()


# ============================================================================
# CONFIG
# ============================================================================
COMP_START = pd.Timestamp("2026-04-01")
COMP_END = pd.Timestamp("2027-03-31")
POINTS_PER = 5_000_000

HIDDEN_COLS = {"Affina_Revenue", "Affina_rate_bonus"}

LEVEL_MAP = {
    "BDD": "Giam Doc", "SD": "Giam Doc", "RMD": "Giam Doc",
    "TSA Manager": "Giam Doc",
    "BDM": "Truong Phong", "SM": "Truong Phong", "RMM": "Truong Phong",
    "TSA Team Leader": "Truong Phong",
    "CTV": "Chuyen Vien", "CVKD": "Chuyen Vien",
    "AG": "Chuyen Vien", "RMC": "Chuyen Vien",
    "CTV TSA": "Chuyen Vien", "TSA": "Chuyen Vien",
}

TOP_N = {"Giam Doc": 3, "Truong Phong": 5, "Chuyen Vien": 5}

QUARTERS = {
    "Q1 (04-06/2026)": (pd.Timestamp("2026-04-01"), pd.Timestamp("2026-06-30")),
    "Q2 (07-09/2026)": (pd.Timestamp("2026-07-01"), pd.Timestamp("2026-09-30")),
    "Q3 (10-12/2026)": (pd.Timestamp("2026-10-01"), pd.Timestamp("2026-12-31")),
    "Q4 (01-03/2027)": (pd.Timestamp("2027-01-01"), pd.Timestamp("2027-03-31")),
}

QUARTER_RANK_BONUS = {1: 10, 2: 5, 3: 3}


# ============================================================================
# HELPERS
# ============================================================================
def _classify_level(chuc_danh: str) -> str:
    if pd.isna(chuc_danh):
        return "Chuyen Vien"
    cd = str(chuc_danh).strip()
    if cd in LEVEL_MAP:
        return LEVEL_MAP[cd]
    cd_upper = cd.upper()
    if "BDD" in cd_upper or "GIAM DOC" in cd_upper:
        return "Giam Doc"
    if "BDM" in cd_upper or "TRUONG" in cd_upper:
        return "Truong Phong"
    return "Chuyen Vien"


def _compute_points(revenue: float) -> int:
    if pd.isna(revenue) or revenue <= 0:
        return 0
    return int(revenue // POINTS_PER)


def _get_current_quarter() -> str:
    today = pd.Timestamp.now().normalize()
    for qname, (qs, qe) in QUARTERS.items():
        if qs <= today <= qe:
            return qname
    return list(QUARTERS.keys())[-1]


def _is_quarter_closed(qname: str) -> bool:
    _, qe = QUARTERS[qname]
    return pd.Timestamp.now().normalize() > qe


def _normalize_phone(val) -> str:
    if pd.isna(val) or str(val).strip() in ("", "nan", "None"):
        return ""
    return str(val).strip().lstrip("0")


def _export_excel(df_export: pd.DataFrame, sheet_name: str = "KPI Competition") -> BytesIO:
    output = BytesIO()
    try:
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        df_export.to_excel(output, index=False, engine="openpyxl")
        output.seek(0)
        return output

    for col in df_export.select_dtypes(include=["datetime64[ns, UTC]", "datetime64[ns]"]).columns:
        df_export[col] = df_export[col].dt.strftime("%d/%m/%Y")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill("solid", fgColor="7038A0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side("thin"), right=Side("thin"),
            top=Side("thin"), bottom=Side("thin")
        )
        center = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

            max_len = len(str(cell.value)) + 2
            for row_idx in range(2, min(ws.max_row + 1, 200)):
                data_cell = ws.cell(row=row_idx, column=col_idx)
                data_cell.border = thin_border
                data_cell.alignment = Alignment(vertical="center")
                val_len = len(str(data_cell.value)) if data_cell.value else 0
                max_len = max(max_len, val_len + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 35)

        top3_fill = PatternFill("solid", fgColor="FDF2FB")
        for row_idx in range(2, min(5, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = top3_fill

    output.seek(0)
    return output


# ============================================================================
# MAIN
# ============================================================================
st.title("KPI Competition — CLB Tinh Hoa Affina")
st.markdown(
    "**Chu ky thi dua:** 01/04/2026 - 31/03/2027  |  "
    "**Giai thuong:** 13 suat du lich Trung Quoc  |  "
    "**Cong bo:** Thang 04/2027"
)

# ============================================================================
# THE LE THI DUA (expander)
# ============================================================================
with st.expander("**THE LE THI DUA — CLB Tinh Hoa Affina 2026-2027**", expanded=False):
    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown("#### Doi tuong thi dua")
        st.markdown("""
| Cap bac | Kenh Core | Kenh Neo | Kenh TSA | So luong giai |
|---------|-----------|----------|----------|--------------|
| Giam Doc PTKD | BDD | SD/RMD | TSA Manager | **03 suat** |
| Truong Phong PTKD | BDM | SM/RMM | TSA Team Leader | **05 suat** |
| Chuyen Vien KD | CVKD | AG/RMC | CTV TSA | **05 suat** |
""")
        st.markdown("""
**Luu y:**
- Cap Truong Phong va Giam Doc co the thi dua cung hang hoac thi dua voi Chuyen Vien KD.
- Moi ca nhan chi nhan toi da **01 ve thuong**. Truong hop dat tieu chuan o 02 nhom thi nhan thuong theo chuc danh cao nhat.
""")

    with col_right:
        st.markdown("#### Tieu chi xet thuong")
        st.markdown("""
**Giam Doc — CORE BDD** (Top 3):
- KPI target: **450,000,000 VND/thang** (So tien thanh toan cua team)
- Dat >= 70% KPI hang thang trong it nhat 03 thang
- Trung binh cac thang >= 50% KPI
- Xep hang theo trung binh KPI %

**Truong Phong — CORE BDM** (Top 5):
- KPI target: **200,000,000 VND/thang** (So tien thanh toan cua team)
- Dat >= 70% KPI hang thang trong it nhat 03 thang
- Trung binh cac thang >= 50% KPI
- Xep hang theo trung binh KPI %

**NEO/TSA** (SM, SD, Teamlead, Manager):
- Xep hang theo tong doanh thu team (khong ap dung % KPI)

**Chuyen Vien KD** (Top 5):
- Top 05 co tong diem quy doi cao nhat
- Moi 5 trieu dong doanh thu ca nhan = **1 diem**
- Moi 5 trieu dong doanh thu tu Nguoi duoc gioi thieu = **1 diem** cho Nguoi gioi thieu
- Tinh toan thuc hien **tung thang** (thang nao tinh thang do)
- Bonus rank **theo quy**: Hang 1 = +10, Hang 2 = +5, Hang 3 = +3
- Tiebreaker: bang diem → ai co tong doanh thu phi cao hon duoc xep tren
""")

    st.markdown("**Thoi gian:** 01/04/2026 - 31/03/2027  |  **Cong bo ket qua:** Thang 04/2027  |  **Trao thuong:** Du kien Quy 2-3/2027")

st.divider()

# ── Load data ──
df_all = load_master_data()
if df_all.empty:
    st.warning("Chua co du lieu.")
    st.stop()

df_all[DATE_COL] = pd.to_datetime(df_all[DATE_COL], errors="coerce")
df = df_all[(df_all[DATE_COL] >= COMP_START) & (df_all[DATE_COL] <= COMP_END)].copy()

if df.empty:
    st.warning(
        f"Chua co du lieu trong chu ky thi dua ({COMP_START.strftime('%d/%m/%Y')} - {COMP_END.strftime('%d/%m/%Y')})."
    )
    st.stop()

for col in HIDDEN_COLS:
    if col in df.columns:
        df = df.drop(columns=[col])

if "Chức danh" in df.columns:
    df["Cap thi dua"] = df["Chức danh"].apply(_classify_level)
else:
    df["Cap thi dua"] = "Chuyen Vien"

df["month"] = df[DATE_COL].dt.to_period("M")

sale_col = "Họ tên sale" if "Họ tên sale" in df.columns else "Họ tên"
if sale_col not in df.columns:
    st.error("Khong tim thay cot ten sale.")
    st.stop()


# ============================================================================
# INLINE FILTERS (in page body, not sidebar only)
# ============================================================================
st.markdown("### Bo loc")

fcol1, fcol2, fcol3, fcol4 = st.columns(4)

with fcol1:
    level_options = ["Tat ca", "Giam Doc", "Truong Phong", "Chuyen Vien"]
    level_filter = st.selectbox("Cap thi dua", options=level_options, index=0, key="kpi_level_inline")

with fcol2:
    if "Source" in df.columns:
        source_options = sorted(df["Source"].dropna().unique().tolist())
        sel_sources = st.multiselect("Source", options=source_options, default=source_options, key="kpi_src_inline")
    else:
        sel_sources = None

with fcol3:
    if "Channel" in df.columns:
        channel_options = sorted(df["Channel"].dropna().unique().tolist())
        sel_channels = st.multiselect("Channel", options=channel_options, default=channel_options, key="kpi_ch_inline")
    else:
        sel_channels = None

with fcol4:
    quarter_options = ["Toan chu ky"] + list(QUARTERS.keys())
    sel_quarter = st.selectbox("Quy", options=quarter_options, index=0, key="kpi_q_inline")

# Apply filters
if level_filter != "Tat ca":
    df = df[df["Cap thi dua"] == level_filter]
if sel_sources is not None and sel_sources:
    df = df[df["Source"].isin(sel_sources)]
if sel_channels is not None and sel_channels:
    df = df[df["Channel"].isin(sel_channels)]
if sel_quarter != "Toan chu ky":
    qs_f, qe_f = QUARTERS[sel_quarter]
    df = df[(df[DATE_COL] >= qs_f) & (df[DATE_COL] <= qe_f)]

# Second row filters
fcol5, fcol6, fcol7, fcol8 = st.columns(4)
with fcol5:
    all_sales_list = sorted(df[sale_col].dropna().unique().tolist())
    sel_sale = st.multiselect("Tim sale", options=all_sales_list, default=[], key="kpi_sale_inline",
                              placeholder="Tat ca sale")
with fcol6:
    if "Loại bảo hiểm" in df.columns:
        lbh_options = sorted(df["Loại bảo hiểm"].dropna().unique().tolist())
        sel_lbh = st.multiselect("Loai BH", options=lbh_options, default=lbh_options, key="kpi_lbh_inline")
    else:
        sel_lbh = None
with fcol7:
    month_options = sorted(df["month"].unique().tolist())
    month_labels = [f"T{m.month:02d}/{m.year}" for m in month_options]
    sel_months = st.multiselect("Thang", options=month_labels, default=[], key="kpi_month_inline",
                                placeholder="Tat ca thang")
with fcol8:
    st.markdown("")
    st.markdown("")

if sel_sale:
    df = df[df[sale_col].isin(sel_sale)]
if sel_lbh is not None and sel_lbh and "Loại bảo hiểm" in df.columns:
    df = df[df["Loại bảo hiểm"].isin(sel_lbh)]
if sel_months:
    selected_periods = [month_options[month_labels.index(m)] for m in sel_months]
    df = df[df["month"].isin(selected_periods)]

if df.empty:
    empty_state("Khong co du lieu sau filter.")
    st.stop()

st.divider()


# ============================================================================
# 1. TIEN DO CHU KY + QUY HIEN TAI
# ============================================================================
st.markdown("### Tien do chu ky thi dua")

today = pd.Timestamp.now().normalize()
days_elapsed = max(0, min((today - COMP_START).days, (COMP_END - COMP_START).days))
days_remaining = max(0, (COMP_END - today).days)
pct_elapsed = days_elapsed / (COMP_END - COMP_START).days * 100
months_elapsed = min(12, max(0, (today.year - COMP_START.year) * 12 + today.month - COMP_START.month))
current_q = _get_current_quarter()

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Ngay da qua", f"{days_elapsed} / {(COMP_END - COMP_START).days}")
c2.metric("Ngay con lai", fmt_num(days_remaining))
c3.metric("Quy hien tai", current_q.split(" ")[0])
c4.metric("Tien do", f"{pct_elapsed:.1f}%")
c5.metric("So sale tham gia", fmt_num(df[sale_col].nunique()))

st.progress(min(pct_elapsed / 100, 1.0))

st.divider()


# ============================================================================
# COMPUTE SCORING (MONTHLY POINTS + QUARTERLY BONUS + REFERRAL)
# ============================================================================

# --- Monthly personal revenue per sale ---
monthly_rev = df.groupby([sale_col, "month"]).agg(
    revenue=("Doanh thu trước thuế", "sum"),
    n_hd=("Số hợp đồng", "nunique") if "Số hợp đồng" in df.columns else (sale_col, "count"),
    cap_thi_dua=("Cap thi dua", "first"),
    source=("Source", "first") if "Source" in df.columns else (sale_col, "first"),
).reset_index()

monthly_rev["diem_ca_nhan"] = monthly_rev["revenue"].apply(_compute_points)

# --- Referral points ---
referral_col = "Người giới thiệu"
has_referral = referral_col in df.columns and df[referral_col].notna().any()

if has_referral:
    df["_ref_phone"] = df[referral_col].apply(_normalize_phone)
    df["_sale_phone"] = df["SĐT sale"].apply(_normalize_phone) if "SĐT sale" in df.columns else ""

    ref_df = df[df["_ref_phone"] != ""].copy()

    if not ref_df.empty and "SĐT sale" in df.columns:
        phone_to_sale = df.drop_duplicates(subset=["_sale_phone"]).set_index("_sale_phone")[sale_col].to_dict()
        phone_to_sale.pop("", None)

        ref_df["referrer_name"] = ref_df["_ref_phone"].map(phone_to_sale)
        ref_matched = ref_df[ref_df["referrer_name"].notna()].copy()

        if not ref_matched.empty:
            referral_monthly = ref_matched.groupby(["referrer_name", "month"]).agg(
                ref_revenue=("Doanh thu trước thuế", "sum")
            ).reset_index()
            referral_monthly.columns = [sale_col, "month", "ref_revenue"]
            referral_monthly["diem_gioi_thieu"] = referral_monthly["ref_revenue"].apply(_compute_points)
        else:
            referral_monthly = pd.DataFrame(columns=[sale_col, "month", "ref_revenue", "diem_gioi_thieu"])
    else:
        referral_monthly = pd.DataFrame(columns=[sale_col, "month", "ref_revenue", "diem_gioi_thieu"])
else:
    referral_monthly = pd.DataFrame(columns=[sale_col, "month", "ref_revenue", "diem_gioi_thieu"])

# Merge referral into monthly
monthly_rev = monthly_rev.merge(
    referral_monthly[[sale_col, "month", "diem_gioi_thieu", "ref_revenue"]],
    on=[sale_col, "month"], how="left"
)
monthly_rev["diem_gioi_thieu"] = monthly_rev["diem_gioi_thieu"].fillna(0).astype(int)
monthly_rev["ref_revenue"] = monthly_rev["ref_revenue"].fillna(0)
monthly_rev["diem_thang"] = monthly_rev["diem_ca_nhan"] + monthly_rev["diem_gioi_thieu"]

# --- Quarter assignment ---
def _month_to_quarter(m):
    ts = m.to_timestamp()
    for qn, (qs, qe) in QUARTERS.items():
        if qs <= ts <= qe:
            return qn
    return ""

monthly_rev["quarter"] = monthly_rev["month"].apply(_month_to_quarter)
monthly_rev["month_ts"] = monthly_rev["month"].dt.to_timestamp()

# --- Quarterly rank bonus (PDF: "điểm quy đổi cao nhất trong quý") ---
quarterly_points = monthly_rev.groupby([sale_col, "quarter"]).agg(
    quy_diem=("diem_thang", "sum")
).reset_index()

quarterly_points["quarter_rank"] = quarterly_points.groupby("quarter")["quy_diem"].rank(
    ascending=False, method="min"
)
quarterly_points["bonus_quy"] = quarterly_points["quarter_rank"].apply(
    lambda r: QUARTER_RANK_BONUS.get(int(r), 0)
)

# ============================================================================
# 2. DIEM THEO THANG & QUY
# ============================================================================
st.markdown("### Diem theo thang va quy")

closed_qs = [q for q in QUARTERS if _is_quarter_closed(q)]

tab_labels = []
for q in QUARTERS:
    if _is_quarter_closed(q):
        tab_labels.append(f"{q} — DA CHOT")
    elif q == current_q:
        tab_labels.append(f"{q} — DANG DIEN RA")
    else:
        tab_labels.append(q)

active_tabs = []
for i, q in enumerate(QUARTERS):
    has_data = monthly_rev[monthly_rev["quarter"] == q].shape[0] > 0
    if _is_quarter_closed(q) or q == current_q or has_data:
        active_tabs.append((tab_labels[i], q))

if active_tabs:
    tabs = st.tabs([t[0] for t in active_tabs])

    for tab, (label, qname) in zip(tabs, active_tabs):
        with tab:
            qs, qe = QUARTERS[qname]
            q_data = monthly_rev[monthly_rev["quarter"] == qname]
            q_bonus = quarterly_points[quarterly_points["quarter"] == qname]
            is_closed = _is_quarter_closed(qname)

            if q_data.empty:
                st.info(f"Chua co du lieu cho {qname}.")
                continue

            if is_closed:
                st.success(f"Quy nay da ket thuc ({qs.strftime('%d/%m')} - {qe.strftime('%d/%m/%Y')})")
            else:
                days_in_q = (qe - qs).days + 1
                days_done = min((today - qs).days, days_in_q)
                st.info(f"Dang dien ra — {days_done}/{days_in_q} ngay ({days_done/days_in_q*100:.0f}%)")

            months_in_q = sorted(q_data["month"].unique())

            q_pivot = q_data.pivot_table(
                index=[sale_col, "cap_thi_dua"],
                columns="month",
                values="diem_thang",
                aggfunc="sum",
                fill_value=0,
            ).reset_index()

            month_cols = [c for c in q_pivot.columns if isinstance(c, pd.Period)]
            q_pivot["Tong diem quy"] = q_pivot[month_cols].sum(axis=1)

            # Add quarterly bonus
            q_pivot = q_pivot.merge(
                q_bonus[[sale_col, "bonus_quy"]],
                on=sale_col, how="left"
            )
            q_pivot["bonus_quy"] = q_pivot["bonus_quy"].fillna(0).astype(int)
            q_pivot["Tong + Bonus"] = q_pivot["Tong diem quy"] + q_pivot["bonus_quy"]

            q_pivot = q_pivot.sort_values("Tong + Bonus", ascending=False).reset_index(drop=True)
            q_pivot.insert(0, "Hang", range(1, len(q_pivot) + 1))

            rename_map = {sale_col: "Ho ten", "cap_thi_dua": "Cap", "bonus_quy": "Bonus QR"}
            for m in month_cols:
                rename_map[m] = f"T{m.month:02d}/{m.year}"
            disp_q = q_pivot.rename(columns=rename_map)

            st.dataframe(disp_q.head(30), hide_index=True, use_container_width=True)

            # Top 3 medals
            col_q1, col_q2, col_q3 = st.columns(3)
            top3_q = q_pivot.head(3)
            medals = ["1.", "2.", "3."]
            cols_q = [col_q1, col_q2, col_q3]
            for i, (_, row) in enumerate(top3_q.iterrows()):
                if i < 3:
                    with cols_q[i]:
                        bonus_txt = f" (+{int(row['bonus_quy'])})" if row["bonus_quy"] > 0 else ""
                        st.metric(
                            f"{medals[i]} {row[sale_col]}",
                            f"{int(row['Tong + Bonus'])} diem{bonus_txt}",
                        )

st.divider()


# ============================================================================
# 3. BANG XEP HANG TONG HOP (toan chu ky) — CHUYEN VIEN
# ============================================================================
st.markdown("### Bang xep hang tong hop — Chuyen Vien KD (toan chu ky)")

# Total personal + referral points per month, summed over all months
total_monthly_points = monthly_rev.groupby(sale_col).agg(
    diem_ca_nhan_total=("diem_ca_nhan", "sum"),
    diem_gioi_thieu_total=("diem_gioi_thieu", "sum"),
    total_revenue=("revenue", "sum"),
    ref_revenue_total=("ref_revenue", "sum"),
    n_months=("month", "nunique"),
    cap_thi_dua=("cap_thi_dua", "first"),
    source=("source", "first"),
).reset_index()

# Total quarterly bonus
total_q_bonus = quarterly_points.groupby(sale_col)["bonus_quy"].sum().reset_index()
total_q_bonus.columns = [sale_col, "bonus_quy_total"]

# Number of quarters in top 3
q_top3 = quarterly_points[quarterly_points["quarter_rank"] <= 3].groupby(sale_col).size().reset_index(name="Quy top 3")

# Total HD
total_hd = df.groupby(sale_col).agg(
    n_hd=("Số hợp đồng", "nunique") if "Số hợp đồng" in df.columns else (sale_col, "count"),
    chuc_danh=("Chức danh", "first") if "Chức danh" in df.columns else (sale_col, "first"),
    channel=("Channel", "first") if "Channel" in df.columns else (sale_col, "first"),
).reset_index()

ranking = total_monthly_points.merge(total_q_bonus, on=sale_col, how="left")
ranking = ranking.merge(q_top3, on=sale_col, how="left")
ranking = ranking.merge(total_hd, on=sale_col, how="left")

ranking["bonus_quy_total"] = ranking["bonus_quy_total"].fillna(0).astype(int)
ranking["Quy top 3"] = ranking["Quy top 3"].fillna(0).astype(int)

ranking["Diem DT"] = ranking["diem_ca_nhan_total"] + ranking["diem_gioi_thieu_total"]
ranking["Tong diem"] = ranking["Diem DT"] + ranking["bonus_quy_total"]

# Tiebreaker: equal points → higher total revenue wins
ranking = ranking.sort_values(
    ["Tong diem", "total_revenue"], ascending=[False, False]
).reset_index(drop=True)
ranking.insert(0, "Hang", range(1, len(ranking) + 1))

# Display tables by level
disp = ranking[[
    "Hang", sale_col, "cap_thi_dua", "source", "channel",
    "n_hd", "total_revenue", "diem_ca_nhan_total", "diem_gioi_thieu_total",
    "Diem DT", "bonus_quy_total", "Tong diem",
    "Quy top 3", "n_months"
]].copy()
disp.columns = [
    "Hang", "Ho ten", "Cap", "Source", "Channel",
    "So HD", "Tong DT", "Diem CN", "Diem GT",
    "Diem DT", "Bonus QR", "Tong diem",
    "Quy top 3", "Thang active"
]
disp["Tong DT"] = disp["Tong DT"].apply(lambda v: fmt_vnd(v, short=True))

tab_all, tab_gd, tab_tp, tab_cv = st.tabs(["Tat ca", "Giam Doc (Top 3)", "Truong Phong (Top 5)", "Chuyen Vien (Top 5)"])

with tab_all:
    st.dataframe(disp, hide_index=True, use_container_width=True, height=450)

with tab_gd:
    gd = disp[disp["Cap"] == "Giam Doc"].reset_index(drop=True)
    gd["Hang"] = range(1, len(gd) + 1)
    if not gd.empty:
        st.dataframe(gd.head(20), hide_index=True, use_container_width=True)
        st.success(f"Vung giai thuong: Top **3** — hien co **{min(3, len(gd))}** nguoi du dieu kien xet")
        st.caption("*Giam Doc CORE xet theo KPI quan ly (muc 4 ben duoi). NEO/TSA xep hang theo DT team.*")
    else:
        empty_state("Khong co Giam Doc trong du lieu.")

with tab_tp:
    tp = disp[disp["Cap"] == "Truong Phong"].reset_index(drop=True)
    tp["Hang"] = range(1, len(tp) + 1)
    if not tp.empty:
        st.dataframe(tp.head(20), hide_index=True, use_container_width=True)
        st.success(f"Vung giai thuong: Top **5** — hien co **{min(5, len(tp))}** nguoi du dieu kien xet")
        st.caption("*Truong Phong CORE xet theo KPI quan ly (muc 4 ben duoi). NEO/TSA xep hang theo DT team.*")
    else:
        empty_state("Khong co Truong Phong.")

with tab_cv:
    cv = disp[disp["Cap"] == "Chuyen Vien"].reset_index(drop=True)
    cv["Hang"] = range(1, len(cv) + 1)
    if not cv.empty:
        st.dataframe(cv.head(30), hide_index=True, use_container_width=True)
        st.success(f"Vung giai thuong: Top **5** — hien co **{min(5, len(cv))}** nguoi du dieu kien xet")
    else:
        empty_state("Khong co Chuyen Vien.")

st.divider()


# ============================================================================
# 4. KPI QUAN LY — GIAM DOC & TRUONG PHONG
# ============================================================================
st.markdown("### KPI quan ly — Giam Doc & Truong Phong")

ql_col_1 = "QUẢN LÝ CẤP 1 (BDM)"
ql_col_2 = "QUẢN LÝ CẤP 2 (BDD)"
kpi_revenue_col = "Số tiền thanh toán"

KPI_TARGETS = {
    "BDM": 200_000_000,
    "BDD": 450_000_000,
}

if ql_col_1 in df.columns or ql_col_2 in df.columns:

    # ── Helper: compute monthly KPI % for a manager column (CORE only) ──
    def _compute_kpi_monthly(data: pd.DataFrame, ql_col: str, target: int) -> pd.DataFrame:
        """Group by manager + month, sum Số tiền thanh toán, compute KPI %."""
        if ql_col not in data.columns or kpi_revenue_col not in data.columns:
            return pd.DataFrame()
        core_data = data[data["Channel"].str.lower().str.strip() == "core"] if "Channel" in data.columns else data
        if core_data.empty:
            return pd.DataFrame()
        monthly = core_data.groupby([ql_col, "month"]).agg(
            team_revenue=(kpi_revenue_col, "sum"),
        ).reset_index()
        monthly = monthly[monthly[ql_col].notna() & (monthly[ql_col].str.strip() != "")]
        monthly["kpi_pct"] = monthly["team_revenue"] / target * 100
        monthly["dat_70"] = monthly["kpi_pct"] >= 70
        return monthly

    def _build_kpi_ranking(monthly_kpi: pd.DataFrame, ql_col: str) -> pd.DataFrame:
        """From monthly KPI data, compute ranking with eligibility conditions."""
        if monthly_kpi.empty:
            return pd.DataFrame()
        summary = monthly_kpi.groupby(ql_col).agg(
            tb_kpi=("kpi_pct", "mean"),
            tong_kpi=("kpi_pct", "sum"),
            so_thang_70=("dat_70", "sum"),
            so_thang=("month", "nunique"),
            tong_dt_team=("team_revenue", "sum"),
        ).reset_index()
        summary["du_dk"] = (summary["so_thang_70"] >= 3) & (summary["tb_kpi"] >= 50)
        summary = summary.sort_values(
            ["du_dk", "tb_kpi", "tong_dt_team"], ascending=[False, False, False]
        ).reset_index(drop=True)
        summary.insert(0, "Hang", range(1, len(summary) + 1))
        return summary

    # ── Compute KPI for CORE BDM (Trưởng Phòng) and BDD (Giám Đốc) ──
    kpi_monthly_tp = _compute_kpi_monthly(df, ql_col_1, KPI_TARGETS["BDM"])
    kpi_monthly_gd = _compute_kpi_monthly(df, ql_col_2, KPI_TARGETS["BDD"])
    kpi_rank_tp = _build_kpi_ranking(kpi_monthly_tp, ql_col_1)
    kpi_rank_gd = _build_kpi_ranking(kpi_monthly_gd, ql_col_2)

    # ── NEO/TSA: rank by total team revenue only ──
    def _team_revenue_ranking(data: pd.DataFrame, ql_col: str, exclude_channel: str = "core") -> pd.DataFrame:
        if ql_col not in data.columns:
            return pd.DataFrame()
        non_core = data[data["Channel"].str.lower().str.strip() != exclude_channel] if "Channel" in data.columns else pd.DataFrame()
        if non_core.empty:
            return pd.DataFrame()
        team = non_core.groupby(ql_col).agg(
            team_revenue=("Doanh thu trước thuế", "sum"),
            n_members=(sale_col, "nunique"),
            n_months=(DATE_COL, lambda x: x.dt.to_period("M").nunique()),
        ).reset_index()
        team = team[team[ql_col].notna() & (team[ql_col].str.strip() != "")]
        team = team.sort_values("team_revenue", ascending=False).reset_index(drop=True)
        team.insert(0, "Hang", range(1, len(team) + 1))
        return team

    neo_tsa_tp = _team_revenue_ranking(df, ql_col_1)
    neo_tsa_gd = _team_revenue_ranking(df, ql_col_2)

    # ── Display ──
    tab_kpi_gd, tab_kpi_tp = st.tabs(["Giam Doc — CORE (BDD)", "Truong Phong — CORE (BDM)"])

    with tab_kpi_tp:
        st.caption(f"KPI target: **{fmt_vnd(KPI_TARGETS['BDM'])}**/thang (So tien thanh toan team) | DK: >=70% x 3 thang, TB >=50%")
        if not kpi_rank_tp.empty:
            disp_tp = kpi_rank_tp.copy()
            disp_tp["tb_kpi"] = disp_tp["tb_kpi"].apply(lambda x: f"{x:.1f}%")
            disp_tp["du_dk"] = disp_tp["du_dk"].map({True: "Dat", False: "Chua"})
            disp_tp["tong_dt_team"] = disp_tp["tong_dt_team"].apply(lambda v: fmt_vnd(v, short=True))
            disp_tp = disp_tp.rename(columns={
                ql_col_1: "Truong Phong", "tb_kpi": "TB KPI %",
                "so_thang_70": "Thang >=70%", "so_thang": "Thang",
                "tong_dt_team": "Tong DT Team", "du_dk": "Du dieu kien",
            })
            disp_tp = disp_tp.drop(columns=["tong_kpi"], errors="ignore")
            st.dataframe(disp_tp, hide_index=True, use_container_width=True)
            n_eligible = int(kpi_rank_tp["du_dk"].sum())
            st.success(f"Vung giai thuong: Top **5** — hien co **{n_eligible}** nguoi du dieu kien")

            # Monthly KPI detail (expandable)
            with st.expander("Chi tiet KPI % theo thang (CORE BDM)"):
                if not kpi_monthly_tp.empty:
                    pivot_tp = kpi_monthly_tp.pivot_table(
                        index=ql_col_1, columns="month", values="kpi_pct", aggfunc="first"
                    ).reset_index()
                    month_cols_tp = [c for c in pivot_tp.columns if isinstance(c, pd.Period)]
                    rename_m = {m: f"T{m.month:02d}/{m.year}" for m in month_cols_tp}
                    pivot_tp = pivot_tp.rename(columns={**rename_m, ql_col_1: "Truong Phong"})
                    for mc in rename_m.values():
                        if mc in pivot_tp.columns:
                            pivot_tp[mc] = pivot_tp[mc].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")
                    st.dataframe(pivot_tp, hide_index=True, use_container_width=True)
        else:
            st.info("Khong co du lieu CORE BDM de tinh KPI.")

        # NEO/TSA Trưởng Phòng — revenue only
        if not neo_tsa_tp.empty:
            st.markdown("---")
            st.markdown("**NEO / TSA — Truong Phong** (xep hang theo DT team, khong tinh % KPI)")
            neo_tsa_tp_disp = neo_tsa_tp.rename(columns={
                ql_col_1: "Truong Phong", "team_revenue": "DT Team",
                "n_members": "So TV", "n_months": "Thang"
            })
            neo_tsa_tp_disp["DT Team"] = neo_tsa_tp_disp["DT Team"].apply(lambda v: fmt_vnd(v, short=True))
            st.dataframe(neo_tsa_tp_disp, hide_index=True, use_container_width=True)

    with tab_kpi_gd:
        st.caption(f"KPI target: **{fmt_vnd(KPI_TARGETS['BDD'])}**/thang (So tien thanh toan team) | DK: >=70% x 3 thang, TB >=50%")
        if not kpi_rank_gd.empty:
            disp_gd = kpi_rank_gd.copy()
            disp_gd["tb_kpi"] = disp_gd["tb_kpi"].apply(lambda x: f"{x:.1f}%")
            disp_gd["du_dk"] = disp_gd["du_dk"].map({True: "Dat", False: "Chua"})
            disp_gd["tong_dt_team"] = disp_gd["tong_dt_team"].apply(lambda v: fmt_vnd(v, short=True))
            disp_gd = disp_gd.rename(columns={
                ql_col_2: "Giam Doc", "tb_kpi": "TB KPI %",
                "so_thang_70": "Thang >=70%", "so_thang": "Thang",
                "tong_dt_team": "Tong DT Team", "du_dk": "Du dieu kien",
            })
            disp_gd = disp_gd.drop(columns=["tong_kpi"], errors="ignore")
            st.dataframe(disp_gd, hide_index=True, use_container_width=True)
            n_eligible_gd = int(kpi_rank_gd["du_dk"].sum())
            st.success(f"Vung giai thuong: Top **3** — hien co **{n_eligible_gd}** nguoi du dieu kien")

            # Monthly KPI detail (expandable)
            with st.expander("Chi tiet KPI % theo thang (CORE BDD)"):
                if not kpi_monthly_gd.empty:
                    pivot_gd = kpi_monthly_gd.pivot_table(
                        index=ql_col_2, columns="month", values="kpi_pct", aggfunc="first"
                    ).reset_index()
                    month_cols_gd = [c for c in pivot_gd.columns if isinstance(c, pd.Period)]
                    rename_m_gd = {m: f"T{m.month:02d}/{m.year}" for m in month_cols_gd}
                    pivot_gd = pivot_gd.rename(columns={**rename_m_gd, ql_col_2: "Giam Doc"})
                    for mc in rename_m_gd.values():
                        if mc in pivot_gd.columns:
                            pivot_gd[mc] = pivot_gd[mc].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")
                    st.dataframe(pivot_gd, hide_index=True, use_container_width=True)
        else:
            st.info("Khong co du lieu CORE BDD de tinh KPI.")

        # NEO/TSA Giám Đốc — revenue only
        if not neo_tsa_gd.empty:
            st.markdown("---")
            st.markdown("**NEO / TSA — Giam Doc** (xep hang theo DT team, khong tinh % KPI)")
            neo_tsa_gd_disp = neo_tsa_gd.rename(columns={
                ql_col_2: "Giam Doc", "team_revenue": "DT Team",
                "n_members": "So TV", "n_months": "Thang"
            })
            neo_tsa_gd_disp["DT Team"] = neo_tsa_gd_disp["DT Team"].apply(lambda v: fmt_vnd(v, short=True))
            st.dataframe(neo_tsa_gd_disp, hide_index=True, use_container_width=True)

else:
    st.info("Khong co du lieu quan ly trong dataset hien tai.")

st.divider()


# ============================================================================
# 5. TOP 10 BAR CHART + KHOANG CACH
# ============================================================================
st.markdown("### Top 10 — Tong diem + Khoang cach")
col_bar, col_gap = st.columns([3, 2])

with col_bar:
    top10 = ranking.head(10)
    if not top10.empty:
        fig = go.Figure()
        fig.add_trace(go.Bar(
            y=top10[sale_col], x=top10["Diem DT"],
            name="Diem DT (CN + GT)", orientation="h",
            marker_color="#B44BC8",
            text=top10["Diem DT"], textposition="inside",
        ))
        fig.add_trace(go.Bar(
            y=top10[sale_col], x=top10["bonus_quy_total"],
            name="Bonus rank quy", orientation="h",
            marker_color="#E85BD8",
            text=top10["bonus_quy_total"], textposition="inside",
        ))
        fig.update_layout(barmode="stack", yaxis=dict(autorange="reversed"))
        st.plotly_chart(apply_plotly_layout(fig, title="Top 10 tong diem (stacked)", height=400),
                        use_container_width=True)

with col_gap:
    st.markdown("**Khoang cach den vung giai thuong**")
    for cap, top_n in TOP_N.items():
        cap_df = ranking[ranking["cap_thi_dua"] == cap].head(top_n + 3)
        if len(cap_df) > top_n:
            threshold = cap_df.iloc[top_n - 1]["Tong diem"]
            first_out = cap_df.iloc[top_n]
            gap = threshold - first_out["Tong diem"]
            gap_revenue = gap * POINTS_PER
            st.markdown(
                f"**{cap}** (Top {top_n}):  \n"
                f"Nguong vao giai: **{int(threshold)} diem**  \n"
                f"Nguoi dau tien ngoai giai cach **{int(gap)} diem** "
                f"(~ {fmt_vnd(gap_revenue, short=True)} DT)"
            )
        elif len(cap_df) > 0:
            st.markdown(f"**{cap}** (Top {top_n}): Chua du nguoi de so sanh")
        st.markdown("")

st.divider()


# ============================================================================
# 6. TIEN TRINH DIEM THEO THANG (line chart)
# ============================================================================
st.markdown("### Tien trinh tich luy diem theo thang")

top10_names = ranking.head(10)[sale_col].tolist()
if top10_names:
    monthly_top = monthly_rev[monthly_rev[sale_col].isin(top10_names)].copy()
    monthly_top = monthly_top.sort_values(["month_ts", sale_col])
    monthly_top["cum_points"] = monthly_top.groupby(sale_col)["diem_thang"].cumsum()

    fig = px.line(
        monthly_top, x="month_ts", y="cum_points", color=sale_col,
        markers=True,
        labels={"month_ts": "Thang", "cum_points": "Tong diem tich luy", sale_col: "Sale"},
    )
    fig.update_xaxes(dtick="M1", tickformat="%m/%Y")
    fig.update_layout(hovermode="x unified")
    st.plotly_chart(apply_plotly_layout(fig, title="", height=420), use_container_width=True)
else:
    empty_state()

st.divider()


# ============================================================================
# 7. HEATMAP RANK THANG
# ============================================================================
st.markdown("### Xep hang theo thang — Ai dan dau moi thang?")

n_show = st.slider("So sale hien thi", min_value=5, max_value=30, value=10, key="kpi_heatmap_n")

# Monthly rank by points
monthly_rev["month_rank"] = monthly_rev.groupby("month")["diem_thang"].rank(
    ascending=False, method="min"
)

top_n_names = ranking.head(n_show)[sale_col].tolist()
if top_n_names:
    heat_data = monthly_rev[monthly_rev[sale_col].isin(top_n_names)].copy()
    heat_data["month_str"] = heat_data["month"].astype(str)
    heat_pivot = heat_data.pivot_table(
        index=sale_col, columns="month_str", values="month_rank", aggfunc="first"
    )
    heat_pivot = heat_pivot.loc[heat_pivot.mean(axis=1).sort_values().index]

    fig = px.imshow(
        heat_pivot.values,
        x=heat_pivot.columns.tolist(),
        y=heat_pivot.index.tolist(),
        aspect="auto",
        color_continuous_scale=["#5FBFA0", "#FDF2FB", "#E8738F"],
        text_auto=".0f",
        labels=dict(color="Hang"),
    )
    fig.update_layout(height=max(300, 35 * n_show))
    st.plotly_chart(apply_plotly_layout(fig, title="Hang moi thang (1 = dan dau, xanh = tot, hong = thap)"),
                    use_container_width=True)

st.divider()


# ============================================================================
# 8. DOANH THU & DIEM GIOI THIEU
# ============================================================================
if has_referral and not referral_monthly.empty:
    st.markdown("### Diem gioi thieu (Referral)")
    st.caption("Diem tu doanh thu cua Nguoi duoc gioi thieu (moi 5 trieu = 1 diem cho Nguoi gioi thieu)")

    ref_summary = referral_monthly.groupby(sale_col).agg(
        total_ref_revenue=("ref_revenue", "sum"),
        total_ref_points=("diem_gioi_thieu", "sum"),
        n_months_ref=("month", "nunique"),
    ).reset_index().sort_values("total_ref_points", ascending=False).reset_index(drop=True)
    ref_summary.insert(0, "Hang", range(1, len(ref_summary) + 1))
    ref_summary_disp = ref_summary.rename(columns={
        sale_col: "Nguoi gioi thieu",
        "total_ref_revenue": "DT referral",
        "total_ref_points": "Diem GT",
        "n_months_ref": "Thang co ref",
    })
    ref_summary_disp["DT referral"] = ref_summary_disp["DT referral"].apply(lambda v: fmt_vnd(v, short=True))
    st.dataframe(ref_summary_disp.head(20), hide_index=True, use_container_width=True)
    st.divider()


# ============================================================================
# 9. SO SANH 1:1
# ============================================================================
st.markdown("### So sanh 1 vs 1")

all_sales = ranking[sale_col].tolist()
if len(all_sales) >= 2:
    col1, col2 = st.columns(2)
    with col1:
        sale_a = st.selectbox("Sale A", options=all_sales, index=0, key="kpi_a")
    with col2:
        default_b = 1 if len(all_sales) > 1 else 0
        sale_b = st.selectbox("Sale B", options=all_sales, index=default_b, key="kpi_b")

    if sale_a and sale_b and sale_a != sale_b:
        row_a = ranking[ranking[sale_col] == sale_a].iloc[0]
        row_b = ranking[ranking[sale_col] == sale_b].iloc[0]

        metrics = ["Tong diem", "Diem DT", "bonus_quy_total", "total_revenue", "n_hd", "Quy top 3"]
        labels = ["Tong diem", "Diem DT", "Bonus QR", "Doanh thu", "So HD", "Quy top 3"]

        c1, c2, c3 = st.columns([2, 1, 2])
        with c1:
            st.markdown(f"**{sale_a}**")
            st.caption(f"Hang {int(row_a['Hang'])} | {row_a['cap_thi_dua']}")
        with c2:
            st.markdown("<div style='text-align:center'>vs</div>", unsafe_allow_html=True)
        with c3:
            st.markdown(f"**{sale_b}**")
            st.caption(f"Hang {int(row_b['Hang'])} | {row_b['cap_thi_dua']}")

        values_a = [float(row_a[m]) for m in metrics]
        values_b = [float(row_b[m]) for m in metrics]
        max_vals = [max(a, b, 1) for a, b in zip(values_a, values_b)]
        norm_a = [a / m for a, m in zip(values_a, max_vals)]
        norm_b = [b / m for b, m in zip(values_b, max_vals)]

        fig = go.Figure()
        fig.add_trace(go.Scatterpolar(
            r=norm_a + [norm_a[0]], theta=labels + [labels[0]],
            fill="toself", name=sale_a,
            fillcolor="rgba(180, 75, 200, 0.2)", line_color="#B44BC8",
        ))
        fig.add_trace(go.Scatterpolar(
            r=norm_b + [norm_b[0]], theta=labels + [labels[0]],
            fill="toself", name=sale_b,
            fillcolor="rgba(240, 110, 194, 0.2)", line_color="#F06EC2",
        ))
        fig.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 1.1])),
            height=380,
        )
        st.plotly_chart(apply_plotly_layout(fig, title=""), use_container_width=True)

        monthly_ab = monthly_rev[monthly_rev[sale_col].isin([sale_a, sale_b])].copy()
        if not monthly_ab.empty:
            fig2 = px.bar(
                monthly_ab, x="month_ts", y="diem_thang", color=sale_col,
                barmode="group",
                color_discrete_map={sale_a: "#B44BC8", sale_b: "#F06EC2"},
                labels={"month_ts": "Thang", "diem_thang": "Diem"},
            )
            fig2.update_xaxes(dtick="M1", tickformat="%m/%Y")
            st.plotly_chart(apply_plotly_layout(fig2, title="Diem theo thang", height=320),
                            use_container_width=True)
    elif sale_a == sale_b:
        st.info("Chon 2 sale khac nhau de so sanh.")

st.divider()


# ============================================================================
# 10. DU BAO
# ============================================================================
st.markdown("### Du bao cuoi chu ky (uoc tinh)")

if months_elapsed > 0 and not ranking.empty:
    ranking["Toc do diem/thang"] = (ranking["Tong diem"] / max(months_elapsed, 1)).round(1)
    ranking["Du bao cuoi ky"] = (ranking["Toc do diem/thang"] * 12).round(0).astype(int)

    forecast_df = ranking.head(15)[[sale_col, "cap_thi_dua", "Tong diem", "Toc do diem/thang", "Du bao cuoi ky"]].copy()
    forecast_df.columns = ["Ho ten", "Cap", "Diem hien tai", "Toc do/thang", "Du bao cuoi ky (12 thang)"]
    forecast_df.insert(0, "Hang", range(1, len(forecast_df) + 1))
    st.dataframe(forecast_df, hide_index=True, use_container_width=True)

    st.caption(
        "Du bao dua tren gia dinh toc do tich diem giu nguyen. "
        "Thuc te co the thay doi do chuong trinh moi, mua vu, thay doi nhan su."
    )
else:
    st.info("Can it nhat 1 thang du lieu trong chu ky thi dua.")

st.divider()


# ============================================================================
# 11. XUAT EXCEL
# ============================================================================
st.markdown("### Xuat report Excel")

export_choice = st.radio(
    "Chon bao cao",
    options=["Bang xep hang tong hop", "Chi tiet diem theo thang", "Diem theo quy"],
    horizontal=True,
    key="kpi_export_choice",
)

if export_choice == "Bang xep hang tong hop":
    export_df = ranking[[
        "Hang", sale_col, "cap_thi_dua", "source", "channel", "n_hd",
        "total_revenue", "diem_ca_nhan_total", "diem_gioi_thieu_total",
        "Diem DT", "bonus_quy_total", "Tong diem", "Quy top 3", "n_months",
    ]].copy()
    export_df.columns = [
        "Hang", "Ho ten", "Cap thi dua", "Source", "Channel", "So HD",
        "Tong doanh thu", "Diem ca nhan", "Diem gioi thieu",
        "Diem DT", "Bonus rank quy", "Tong diem", "Quy top 3", "Thang active",
    ]
    sheet = "BXH Tong hop"
elif export_choice == "Chi tiet diem theo thang":
    export_df = monthly_rev[[
        sale_col, "month", "revenue", "diem_ca_nhan", "diem_gioi_thieu",
        "diem_thang", "cap_thi_dua", "source"
    ]].copy()
    export_df["month"] = export_df["month"].astype(str)
    export_df.columns = [
        "Ho ten", "Thang", "Doanh thu", "Diem CN", "Diem GT",
        "Diem thang", "Cap thi dua", "Source"
    ]
    export_df = export_df.sort_values(["Thang", "Diem thang"], ascending=[True, False])
    sheet = "Diem theo thang"
else:
    export_df = quarterly_points.copy()
    export_df.columns = ["Ho ten", "Quy", "Diem quy", "Hang quy", "Bonus quy"]
    export_df = export_df.sort_values(["Quy", "Hang quy"])
    sheet = "Diem theo quy"

excel_data = _export_excel(export_df, sheet_name=sheet)
st.download_button(
    label=f"Tai Excel — {export_choice}",
    data=excel_data,
    file_name=f"KPI_Competition_{sheet.replace(' ', '_')}_{today.strftime('%Y%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    key="kpi_download_excel",
)

st.divider()


# ============================================================================
# FOOTER
# ============================================================================
st.markdown(
    "---\n"
    "*Du lieu cap nhat hang ngay. Ket qua chinh thuc do Ban To chuc cong bo thang 04/2027. "
    "Bang xep hang nay chi mang tinh chat tham khao.*"
)
