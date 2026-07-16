"""
================================================================================
 TRANG 5 — CHI TIẾT & EXPORT NÂNG CAO
================================================================================
Cho phép:
  1. Filter data theo bất kỳ tiêu chí nào (sidebar + advanced filter)
  2. Chọn template có sẵn (Financial / Sales / Customer / Full) hoặc custom columns
  3. Chia file thành nhiều sheet (theo Source / Loại BH / Nhà BH / tất cả 1 sheet)
  4. Thêm sheet Summary (KPI tổng hợp)
  5. Format Excel chuyên nghiệp: header đậm, filter, freeze, VNĐ format, cột tự width
  6. Preview trước khi tải
================================================================================
"""
from io import BytesIO
from datetime import datetime

import pandas as pd
import streamlit as st

from lib.data import (
    DATE_COL,
    apply_filters, empty_state,
    fmt_num, fmt_vnd,
    load_master_data, render_sidebar_filters,
)

st.set_page_config(page_title="Chi tiết & Export", layout="wide")

# ============================================================================
# EXCEL FORMATTING
# ============================================================================
try:
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Cột nào là tiền VNĐ (auto format)
CURRENCY_COLS = {
    "Số tiền thanh toán", "Phí BH (VNĐ)", "Doanh thu trước thuế",
    "EST_Bonus", "Affina_Revenue", "Thưởng Teamlead", "Incentive OVE",
    "SM_OR", "SD_OR", "SM_IO", "SD_IO", "BDM_bonus", "BDD_bonus",
    "Chi Agency", "Chi QL", "Budget Neo T6",
}
DATE_COLS = {
    "Ngày thanh toán", "Ngày bắt đầu", "Ngày kết thúc",
    "Ngày sinh NĐBH", "Ngày sinh NMBH", "_ingested_at",
}


def _apply_sheet_format(ws, df: pd.DataFrame) -> None:
    """Style header + filter + freeze + auto-width + number format."""
    if ws.max_row == 0 or len(df) == 0:
        return

    header_fill = PatternFill(start_color="7D2E78", end_color="7D2E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D4B8D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Data rows — border + number format
    for i, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(i)

        # Number format cho cột tiền / ngày
        if col_name in CURRENCY_COLS:
            fmt = '#,##0" ₫"'
        elif col_name in DATE_COLS or "Ngày" in col_name:
            fmt = "dd/mm/yyyy"
        else:
            fmt = None

        # Apply
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=i)
            cell.border = border
            if fmt:
                cell.number_format = fmt

        # Auto width (approximate — dựa 50 dòng đầu để nhanh)
        try:
            sample_lens = [len(str(v)) for v in df[col_name].head(50).astype(str).tolist()]
            max_len = max([len(str(col_name))] + sample_lens)
            ws.column_dimensions[col_letter].width = min(35, max(10, max_len + 2))
        except Exception:
            ws.column_dimensions[col_letter].width = 15

    # Auto-filter + freeze row 1
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Row height cho header
    ws.row_dimensions[1].height = 32


def _build_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build 1 df tổng hợp KPI để đưa vào sheet Summary."""
    rows = []
    total_rev = df["Doanh thu trước thuế"].sum() if "Doanh thu trước thuế" in df else 0
    n_hd = df["Số hợp đồng"].nunique() if "Số hợp đồng" in df else 0
    n_sale = df["Họ tên sale"].nunique() if "Họ tên sale" in df else 0

    rows.append({"Chỉ số": " Snapshot generated", "Giá trị": pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")})
    rows.append({"Chỉ số": " Tổng số dòng", "Giá trị": len(df)})
    if DATE_COL in df.columns and df[DATE_COL].notna().any():
        rows.append({"Chỉ số": " Từ ngày", "Giá trị": df[DATE_COL].min().strftime("%d/%m/%Y")})
        rows.append({"Chỉ số": " Đến ngày", "Giá trị": df[DATE_COL].max().strftime("%d/%m/%Y")})
    rows.append({"Chỉ số": "─" * 30, "Giá trị": "─" * 30})
    rows.append({"Chỉ số": " Tổng doanh thu (trước thuế)", "Giá trị": f"{total_rev:,.0f} ₫"})
    rows.append({"Chỉ số": " Số HĐ (unique)", "Giá trị": f"{n_hd:,}"})
    rows.append({"Chỉ số": " Số Sale (unique)", "Giá trị": f"{n_sale:,}"})
    if n_hd > 0:
        rows.append({"Chỉ số": " AVG doanh thu / HĐ", "Giá trị": f"{total_rev/n_hd:,.0f} ₫"})

    # Break down by Source
    if "Source" in df.columns:
        rows.append({"Chỉ số": "─" * 30, "Giá trị": "─" * 30})
        rows.append({"Chỉ số": "── BREAKDOWN BY SOURCE ──", "Giá trị": ""})
        for src, grp in df.groupby("Source"):
            rev = grp["Doanh thu trước thuế"].sum() if "Doanh thu trước thuế" in grp else 0
            rows.append({
                "Chỉ số": f"{src}",
                "Giá trị": f"{rev:,.0f} ₫ ({len(grp):,} dòng, {grp['Số hợp đồng'].nunique() if 'Số hợp đồng' in grp else 0:,} HĐ)"
            })

    # Break down by Loại BH
    if "Loại bảo hiểm" in df.columns:
        rows.append({"Chỉ số": "─" * 30, "Giá trị": "─" * 30})
        rows.append({"Chỉ số": "── BREAKDOWN BY LOẠI BH ──", "Giá trị": ""})
        for lb, grp in df.groupby("Loại bảo hiểm"):
            rev = grp["Doanh thu trước thuế"].sum() if "Doanh thu trước thuế" in grp else 0
            rows.append({
                "Chỉ số": f"{lb}",
                "Giá trị": f"{rev:,.0f} ₫ ({len(grp):,} dòng)"
            })

    return pd.DataFrame(rows)


def _sanitize_sheet_name(name: str) -> str:
    """Excel giới hạn 31 ký tự, không có: / \\ ? * [ ]"""
    for ch in ["/", "\\", "?", "*", "[", "]", ":"]:
        name = str(name).replace(ch, "_")
    return name[:31]


def _strip_tz(df: pd.DataFrame) -> pd.DataFrame:
    """Bỏ timezone khỏi mọi cột datetime — Excel không hỗ trợ tz-aware datetime.
    Cột _ingested_at (TIMESTAMPTZ) từ Supabase gây lỗi nếu không strip."""
    df = df.copy()
    for c in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[c]):
            try:
                # Nếu có timezone convert UTC rồi bỏ tz
                if getattr(df[c].dt, "tz", None) is not None:
                    df[c] = df[c].dt.tz_localize(None)
            except (TypeError, AttributeError):
                pass
    return df


def build_excel(
    df: pd.DataFrame,
    columns: list[str],
    sheet_mode: str,
    include_summary: bool,
    file_label: str,
) -> bytes:
    """Sinh file Excel. sheet_mode: 'single' | 'source' | 'loai_bh' | 'nha_bh'"""
    df_out = df[columns].copy()

    # Convert datetime pandas datetime cho openpyxl format đúng
    for c in df_out.columns:
        if c in DATE_COLS or "Ngày" in c:
            df_out[c] = pd.to_datetime(df_out[c], errors="coerce")

    # Strip timezone (Excel không hỗ trợ tz-aware datetime)
    df_out = _strip_tz(df_out)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Summary sheet (đưa lên đầu)
        if include_summary:
            summary_df = _build_summary_df(df)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            if HAS_OPENPYXL:
                ws = writer.sheets["Summary"]
                ws.column_dimensions["A"].width = 40
                ws.column_dimensions["B"].width = 45
                # Style header
                header_fill = PatternFill(start_color="7D2E78", end_color="7D2E78", fill_type="solid")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="FFFFFF", size=11)

        # Data sheets
        if sheet_mode == "single":
            df_out.to_excel(writer, sheet_name="Data", index=False)
            if HAS_OPENPYXL:
                _apply_sheet_format(writer.sheets["Data"], df_out)

        elif sheet_mode == "source" and "Source" in df.columns:
            for src in sorted(df["Source"].dropna().unique()):
                sub = df[df["Source"] == src][columns].copy()
                for c in sub.columns:
                    if c in DATE_COLS or "Ngày" in c:
                        sub[c] = pd.to_datetime(sub[c], errors="coerce")
                sub = _strip_tz(sub)
                sheet_name = _sanitize_sheet_name(src)
                sub.to_excel(writer, sheet_name=sheet_name, index=False)
                if HAS_OPENPYXL:
                    _apply_sheet_format(writer.sheets[sheet_name], sub)

        elif sheet_mode == "loai_bh" and "Loại bảo hiểm" in df.columns:
            for lb in sorted(df["Loại bảo hiểm"].dropna().unique()):
                sub = df[df["Loại bảo hiểm"] == lb][columns].copy()
                for c in sub.columns:
                    if c in DATE_COLS or "Ngày" in c:
                        sub[c] = pd.to_datetime(sub[c], errors="coerce")
                sub = _strip_tz(sub)
                sheet_name = _sanitize_sheet_name(lb)
                sub.to_excel(writer, sheet_name=sheet_name, index=False)
                if HAS_OPENPYXL:
                    _apply_sheet_format(writer.sheets[sheet_name], sub)

        elif sheet_mode == "nha_bh" and "Nhà BH" in df.columns:
            # Chỉ top 15 nhà BH lớn nhất - tránh quá nhiều sheet
            top_partners = (df.groupby("Nhà BH")["Doanh thu trước thuế"].sum()
                              .nlargest(15).index.tolist())
            for p in top_partners:
                sub = df[df["Nhà BH"] == p][columns].copy()
                for c in sub.columns:
                    if c in DATE_COLS or "Ngày" in c:
                        sub[c] = pd.to_datetime(sub[c], errors="coerce")
                sub = _strip_tz(sub)
                sheet_name = _sanitize_sheet_name(p)
                sub.to_excel(writer, sheet_name=sheet_name, index=False)
                if HAS_OPENPYXL:
                    _apply_sheet_format(writer.sheets[sheet_name], sub)

    return output.getvalue()


# ============================================================================
# COLUMN TEMPLATES
# ============================================================================
TEMPLATES = {
    " Full data (tất cả cột)": None, # None = tất cả

    " Financial view": [
        "Ngày thanh toán", "Source", "Channel", "Số hợp đồng", "Loại bảo hiểm",
        "Sản phẩm", "Nhà BH", "Số tiền thanh toán", "Phí BH (VNĐ)",
        "Doanh thu trước thuế", "EST_Bonus", "Affina_Revenue", "Thưởng Teamlead",
    ],

    "Sales view": [
        "Ngày thanh toán", "Source", "Channel", "Số hợp đồng",
        "Họ tên sale", "Chức danh", "SĐT sale",
        "QUẢN LÝ CẤP 1 (BDM)", "QUẢN LÝ CẤP 2 (BDD)", "Quản lý Cấp 3 (BDH)",
        "Sản phẩm", "Nhà BH", "Doanh thu trước thuế", "EST_Bonus",
    ],

    "Customer view": [
        "Ngày thanh toán", "Số hợp đồng", "Loại bảo hiểm", "Sản phẩm",
        "Tên NĐBH", "Ngày sinh NĐBH", "Giới tính NNBH", "CCCD NĐBH",
        "Tên NMBH", "Ngày sinh NMBH", "CCCD NMBH", "Quan hệ",
        "SĐT NMBH", "Email NMBH", "Địa chỉ NMBH",
        "Ngày bắt đầu", "Ngày kết thúc", "Số tiền thanh toán",
    ],

    "Compact view (10 cột)": [
        "Ngày thanh toán", "Source", "Channel", "Số hợp đồng",
        "Họ tên sale", "Loại bảo hiểm", "Sản phẩm", "Nhà BH",
        "Số tiền thanh toán", "Doanh thu trước thuế",
    ],

    "Renewal view (HĐ sắp hết hạn)": [
        "Số hợp đồng", "Loại bảo hiểm", "Sản phẩm", "Nhà BH",
        "Tên NĐBH", "SĐT NMBH", "Email NMBH",
        "Ngày bắt đầu", "Ngày kết thúc",
        "Họ tên sale", "SĐT sale", "Doanh thu trước thuế",
    ],

    "Custom (tự chọn cột)": [], # empty - user picks
}


# ============================================================================
# MAIN
# ============================================================================
st.title("Chi tiết & Export nâng cao")
st.caption(
    "Filter data theo ý muốn — chọn template hoặc custom cột — "
    "xuất Excel có format chuyên nghiệp."
)

df_all = load_master_data()
if df_all.empty:
    st.warning("Chưa có dữ liệu.")
    st.stop()

# Sidebar filters + apply
filters = render_sidebar_filters(df_all)
df = apply_filters(df_all, filters)

if df.empty:
    empty_state("Không có dữ liệu phù hợp filter. Chỉnh lại sidebar.")
    st.stop()

# ============================================================================
# ADVANCED FILTER (ngoài sidebar)
# ============================================================================
with st.expander("Bộ lọc nâng cao (tùy chọn)", expanded=False):
    col1, col2 = st.columns(2)

    with col1:
        # Text search
        search_text = st.text_input(
            "Tìm theo tên khách/HĐ/sale (không phân biệt hoa thường)",
            placeholder="VD: Nguyễn, HD_000123, TRẦN THỊ...",
        )
        if search_text:
            search_lower = search_text.lower()
            search_cols = ["Tên NĐBH", "Tên NMBH", "Số hợp đồng", "Họ tên sale", "SĐT NMBH"]
            search_cols = [c for c in search_cols if c in df.columns]
            if search_cols:
                mask = pd.Series(False, index=df.index)
                for c in search_cols:
                    mask |= df[c].astype(str).str.lower().str.contains(search_lower, na=False)
                df = df[mask]

    with col2:
        # Số tiền thanh toán range
        if "Số tiền thanh toán" in df.columns and len(df) > 0:
            min_v = float(df["Số tiền thanh toán"].min() or 0)
            max_v = float(df["Số tiền thanh toán"].max() or 1)
            if max_v > min_v:
                amt_range = st.slider(
                    "Số tiền thanh toán (VNĐ)",
                    min_value=int(min_v),
                    max_value=int(max_v),
                    value=(int(min_v), int(max_v)),
                    step=max(1, int((max_v - min_v) / 100)),
                    format="%d",
                )
                df = df[(df["Số tiền thanh toán"] >= amt_range[0]) &
                        (df["Số tiền thanh toán"] <= amt_range[1])]

    # BDM/BDD selector
    col3, col4, col5 = st.columns(3)
    with col3:
        if "QUẢN LÝ CẤP 1 (BDM)" in df.columns:
            bdm_opts = sorted(df["QUẢN LÝ CẤP 1 (BDM)"].dropna().unique())
            bdm_sel = st.multiselect("BDM (Cấp 1)", options=bdm_opts, default=[])
            if bdm_sel:
                df = df[df["QUẢN LÝ CẤP 1 (BDM)"].isin(bdm_sel)]
    with col4:
        if "QUẢN LÝ CẤP 2 (BDD)" in df.columns:
            bdd_opts = sorted(df["QUẢN LÝ CẤP 2 (BDD)"].dropna().unique())
            bdd_sel = st.multiselect("BDD (Cấp 2)", options=bdd_opts, default=[])
            if bdd_sel:
                df = df[df["QUẢN LÝ CẤP 2 (BDD)"].isin(bdd_sel)]
    with col5:
        if "Sản phẩm" in df.columns:
            prod_opts = sorted(df["Sản phẩm"].dropna().unique())
            prod_sel = st.multiselect("Sản phẩm", options=prod_opts, default=[])
            if prod_sel:
                df = df[df["Sản phẩm"].isin(prod_sel)]

if df.empty:
    empty_state("Bộ lọc nâng cao loại trừ hết dữ liệu.")
    st.stop()

# ============================================================================
# STATS SAU KHI FILTER
# ============================================================================
st.markdown("### Kết quả filter")
c1, c2, c3, c4 = st.columns(4)
c1.metric("Số dòng", fmt_num(len(df)))
c2.metric("Số HĐ", fmt_num(df["Số hợp đồng"].nunique() if "Số hợp đồng" in df else 0))
c3.metric("Số Sale", fmt_num(df["Họ tên sale"].nunique() if "Họ tên sale" in df else 0))
c4.metric("Tổng doanh thu",
          fmt_vnd(df["Doanh thu trước thuế"].sum() if "Doanh thu trước thuế" in df else 0, short=True))

st.divider()

# ============================================================================
# EXPORT BUILDER
# ============================================================================
st.markdown("### Export Builder — Xuất file Excel")

col_left, col_right = st.columns([2, 1])

with col_left:
    st.markdown("#### 1⃣ Chọn template hoặc custom cột")
    template_name = st.selectbox(
        "Template",
        options=list(TEMPLATES.keys()),
        index=0,
        help="Chọn template có sẵn hoặc chọn 'Custom' để tự pick",
    )

    # Determine default columns
    template_cols = TEMPLATES[template_name]
    if template_cols is None: # Full data
        default_cols = list(df.columns)
    elif template_cols == []: # Custom - empty
        default_cols = list(df.columns[:10]) # 10 cột đầu
    else:
        default_cols = [c for c in template_cols if c in df.columns]

    # Column picker
    available_cols = list(df.columns)
    selected_cols = st.multiselect(
        f"Cột sẽ xuất ({len(default_cols)} cột default)",
        options=available_cols,
        default=default_cols,
        help="Có thể xóa/thêm cột theo ý",
    )

with col_right:
    st.markdown("#### 2⃣ Cấu trúc file")
    sheet_mode_display = st.radio(
        "Chia thành nhiều sheet?",
        options=[
            "1 sheet (tất cả data)",
            "Split by Source (Core/Neo/TSA)",
            "Split by Loại BH (7 sheets)",
            "Split by Top 15 Nhà BH",
        ],
        index=0,
    )
    sheet_mode_map = {
        "1 sheet (tất cả data)": "single",
        "Split by Source (Core/Neo/TSA)": "source",
        "Split by Loại BH (7 sheets)": "loai_bh",
        "Split by Top 15 Nhà BH": "nha_bh",
    }
    sheet_mode = sheet_mode_map[sheet_mode_display]

    include_summary = st.checkbox("Thêm sheet Summary", value=True,
                                   help="Sheet đầu tiên chứa KPI tổng hợp")

# ============================================================================
# PREVIEW
# ============================================================================
st.markdown("#### 3⃣ Preview (20 dòng đầu)")
if selected_cols:
    st.dataframe(
        df[selected_cols].head(20),
        use_container_width=True,
        height=350,
        hide_index=True,
    )

    # Estimate file size
    n_rows = len(df)
    n_cols = len(selected_cols)
    # Rough estimate: ~100 bytes / cell
    est_size_kb = (n_rows * n_cols * 100) / 1024
    if est_size_kb < 1024:
        size_str = f"~{est_size_kb:.0f} KB"
    else:
        size_str = f"~{est_size_kb/1024:.1f} MB"

    st.caption(
        f"File sẽ chứa **{n_rows:,} dòng × {n_cols} cột** — kích thước ước tính: **{size_str}**"
    )

    if n_rows > 100_000:
        st.warning(
            f"Data lớn ({n_rows:,} dòng). Tạo file có thể mất 20-60 giây. "
            f"Cân nhắc thu hẹp khoảng thời gian filter trước."
        )
else:
    st.warning("Chưa chọn cột nào để xuất.")

st.divider()

# ============================================================================
# DOWNLOAD BUTTON
# ============================================================================
st.markdown("#### 4⃣ Tải xuống")

col_dl1, col_dl2 = st.columns([1, 2])

# CSV nhanh (backup)
with col_dl1:
    if selected_cols:
        csv_data = df[selected_cols].to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "Tải CSV (nhanh)",
            data=csv_data,
            file_name=f"affina_data_{datetime.now():%Y%m%d_%H%M}.csv",
            mime="text/csv",
            use_container_width=True,
            help="CSV nhẹ, mở nhanh. Không có format tiền/ngày đẹp như Excel.",
        )

# Excel với format
with col_dl2:
    if selected_cols:
        if not HAS_OPENPYXL:
            st.error(
                "Thiếu `openpyxl`. Thêm vào requirements.txt và deploy lại."
            )
        else:
            # Generate Excel on-demand khi click
            filename_suffix = {
                "single": "single",
                "source": "by_source",
                "loai_bh": "by_type",
                "nha_bh": "by_partner",
            }[sheet_mode]

            file_label = f"affina_{filename_suffix}_{datetime.now():%Y%m%d_%H%M}.xlsx"

            with st.spinner("Đang tạo file Excel với format đẹp..."):
                try:
                    excel_bytes = build_excel(
                        df=df,
                        columns=selected_cols,
                        sheet_mode=sheet_mode,
                        include_summary=include_summary,
                        file_label=file_label,
                    )
                    st.download_button(
                        f"Tải Excel (format đẹp — {len(selected_cols)} cột)",
                        data=excel_bytes,
                        file_name=file_label,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                    st.success(f"File Excel sẵn sàng! Kích thước thực: **{len(excel_bytes)/1024:.0f} KB**")
                except Exception as e:
                    st.error(f"Lỗi tạo Excel: {e}")

st.divider()

# ============================================================================
# HELP
# ============================================================================
with st.expander("ℹ Hướng dẫn — Excel file có gì?"):
    st.markdown("""
**File Excel xuất ra sẽ có:**

- **Header đậm** — chữ trắng trên nền xanh, dễ đọc
- **Auto-filter** — click mũi tên header để lọc trong Excel
- **Freeze row 1** — cuộn xuống vẫn thấy header
- **Cột tiền tự format** — hiển thị `1,234,567 ₫`
- **Cột ngày tự format** — hiển thị `dd/mm/yyyy`
- **Auto-width cột** — không cần chỉnh lại
- **Border thin** — bảng gọn gàng in ra đẹp

**Nếu chọn Summary sheet:**
- Số dòng, khoảng thời gian, tổng doanh thu
- Breakdown theo Source (Core/Neo/TSA)
- Breakdown theo 7 loại BH

**Nếu chọn Split by Source:**
- Sheet `Core` — tất cả HĐ nhánh Core
- Sheet `Neo`
- Sheet `TSA`

**Nếu chọn Split by Loại BH:**
- 7 sheet: `BHSK`, `BHXM`, `BHYT`, `BHOTO`, `BHDL`, `TNDS`, `BHRR`

**Tips:**
- Dùng **Renewal view** template + filter Ngày kết thúc trong 30 ngày tới export cho team CSKH gọi tái tục
- Dùng **Customer view** + sheet mode Split by Nhà BH gửi cho partner báo cáo tháng
- Dùng **Sales view** gửi cho BDM/BDD báo cáo team
    """)
